import base64
import json
import re
import uuid
import time
from smtplib import SMTPAuthenticationError
import subprocess
from django.utils import timezone

from .sort_and_filter_by_cols import get_sort_and_filter_by_cols
import ckeditor
import requests
from rest_framework.exceptions import AuthenticationFailed, ParseError
from mimetypes import guess_type

import openai
from openai import OpenAI

import pytz
from ckeditor.fields import RichTextField
from django.core.files.storage import FileSystemStorage
from django.core.mail import EmailMessage, get_connection, send_mail
from django.core.validators import validate_email
from django.db import IntegrityError
from django.db.models import Q, Sum
from django.db import IntegrityError, models

from django.http import HttpResponse
from django.template.loader import render_to_string
from obfuskey import Obfuskey, alphabets
import logging
import hashlib
from datetime import datetime, timedelta, date, time
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
import csv

from rest_framework import status
from rest_framework.exceptions import ValidationError, ErrorDetail
from django.core.exceptions import ValidationError as DjangoValidationError

from hrms import settings

handle_users = None
from mirage import fields
import traceback
from mirage.fields import EncryptedMixin

em = EncryptedMixin()
tz = pytz.timezone(settings.TIME_ZONE)
openai.api_key = settings.OPENAI_API_KEY
gpt = None
hash_list = {'user': "0123456789aBcDeFhH",
             'cd': "0123456789abCDeFhH",
             'ob': "0123456789ABcDeFhH",
             'doc': "0123456789ABCDeFhH",
             'ojp': "0123456789ABCDEFhH",
             "attendance": "0123456789ABCDEFHh"}

logger = logging.getLogger(__name__)


def get_decrypted_name(value):
    if value is None:
        return None
    return em.crypto.decrypt(value)


def get_decrypted_value(value):
    if value is None:
        return None
    return em.crypto.decrypt(value)


class EncryptedCField(fields.EncryptedCharField):
    def __init__(self, **kwargs):
        super(EncryptedCField, self).__init__(**kwargs)

    def from_db_value(self, value, expression, connection):
        return value


class EncryptedIntField(fields.EncryptedIntegerField):
    def __init__(self, **kwargs):
        super(EncryptedIntField, self).__init__(**kwargs)

    def from_db_value(self, value, expression, connection):
        return value


def get_obfuscator(name, length):
    obfuscator_18 = Obfuskey(hash_list[name], key_length=length)
    return obfuscator_18


def hash_value(val, name, mini=False):
    try:
        size = 11
        # print("val", val)
        obfuscator = get_obfuscator(name, size)
        return obfuscator.get_key(int(val))
    except Exception as e:
        return None


def unhash_value(val, name, mini=False):
    try:
        size = 11
        obfuscator = get_obfuscator(name, size)
        return obfuscator.get_value(val)
    except Exception as e:
        logger.info("exception at unhash_value {0}".format(traceback.format_exc(0)))
        return None


def return_error_response(serializer, raise_exception=False):
    error_list = ""
    count = 0
    for error in serializer.errors:
        if count > 0:
            if error == 'non_field_errors':
                error_list = error_list + ", " + serializer.errors[error][0].title()
            else:
                error_list = error_list + ", " + error.replace("_", " ") + ": " + serializer.errors[error][0].title()
        else:
            if error == 'non_field_errors':
                error_list = serializer.errors[error][0].title()
            else:
                error_list = error.replace("_", " ") + ": " + serializer.errors[error][0].title()
        count += 1
    if raise_exception:
        logger.info("Exception : {0}".format(error_list))
        raise ValueError(error_list)
    return HttpResponse(json.dumps({"error": error_list}), status=status.HTTP_400_BAD_REQUEST)


def update_request_params(request, **kwargs):
    request.query_params._mutable = True

    for key, value in kwargs.items():
        request.query_params[key] = value
    request.query_params._mutable = False

    return request


def update_request(data, **kwargs):
    try:
        data._mutable = True
    except:
        pass
    for key, value in kwargs.items():
        data[key] = value
    try:
        data._mutable = False
    except:
        pass
    return data


def validate_number_with_fixed_size(value, field, size, check_size=True):
    if len(str(value)) != size and check_size:
        logger.info(f"mobile {value} {len(str(value))}")
        raise ValueError("{0} must be of length {1} digits".format(field, size))
    if not str(value).isnumeric():
        raise ValueError("{0} should be a number".format(field))


def hash_file(file):
    eduhash = hashlib.md5()
    eduhash.update(file.read())
    return eduhash.hexdigest()


def form_module_url(doc_id, module):
    return "/".join([module, 'document', str(hash_value(doc_id, 'doc'))])


def document_response(Document, doc_id):
    doc_id = unhash_value(doc_id, 'doc')
    try:
        document = Document.objects.get(id=doc_id)
    except Exception as e:
        return HttpResponse(json.dumps({'error': 'document not found'}), status=status.HTTP_404_NOT_FOUND)
    mimetype, encoding = guess_type(document.file.name, strict=True)
    # TODO check if file exists and proper error message if not
    return HttpResponse(document.file, content_type=mimetype)


def validate_files(filesdata, module=None):
    if module and module in ["appointment_letter", "offer_letter"]:
        allowed_extensions = ["pdf"]
    elif module and module in ["interview_test"]:
        allowed_extensions = ["xls", "xlsx"]
    elif module and module == "onboard":
        allowed_extensions = settings.ONBOARD_ALLOWED_EXTENSIONS
    else:
        allowed_extensions = settings.DEFAULT_ALLOWED_EXTENSIONS
    for filename in filesdata:
        for file in filesdata.getlist(filename, []):
            if file.size > 2 * 1024 * 1024:
                raise ValueError("The maximum allowed file size is 2mb")
            ext = file.name.split(".")
            if not len(ext) > 1:
                raise ValueError(
                    "Please upload a valid file, {0} is not a valid file with an extension".format(filename))
            ext = ext[-1]
            if ext not in allowed_extensions:
                raise ValueError(
                    "{0} is not a valid extension for {1}, allowed extension are {2}".format(ext, filename,
                                                                                             allowed_extensions))


def validate_input_date(date_string):
    try:
        # Attempt to parse the string as a date using the datetime.strptime() method
        if date_string is None:
            return False
        datetime.strptime(date_string, '%Y-%m-%d')
        return True
    except ValueError:
        # If the string is not a valid date, catch the ValueError exception and return False
        return False


def get_formatted_name(profile):
    if not profile:
        return None
    if not profile.middle_name:
        name = profile.first_name or "" + " " + profile.last_name or ""
    else:
        name = profile.first_name or "" + " " + profile.middle_name or "" + " " + profile.last_name or ""
    return name.strip()


def get_full_name(data):
    first_name = data.get("first_name")
    middle_name = data.get("middle_name")
    last_name = data.get("last_name")
    if not middle_name:
        full_name = " ".join([first_name or "", last_name or ""])
    else:
        full_name = " ".join([first_name or "", middle_name or "", last_name or ""])
    return full_name.strip()


def get_dates_range(start_date, end_date):
    t_start_d = start_date
    days = []
    while t_start_d <= end_date:
        days.append(t_start_d)
        t_start_d += timedelta(days=1)
    return days


def get_team_ids(usr):
    # all_team = list(Team.objects.filter(manager=usr).values_list("id", flat=True))

    all_team = list(usr.my_team.filter(is_active=True).values_list("id", flat=True))
    if len(all_team) == 0:
        return None
    return all_team


def get_emp_by_emp_id(emp_id):
    from mapping.models import Profile
    try:
        emp = Profile.objects.filter(emp_id=emp_id).last()
    except Exception as e:
        raise ValueError("invalid emp-id, profile not found")
    return emp


def get_team_by_id(team_id):
    if team_id is None or str(team_id).strip() == "":
        raise ValueError("Please provide the team-id information")
    from team.models import Team
    team = Team.objects.filter(id=team_id).last()
    if team is None:
        raise ValueError("Invalid team-id is provided")
    return team


def get_last_day_of_cur_month_and_first_day_of_next_month(input_date=datetime.today().date()):
    next_28_days = input_date + timedelta(days=28)
    if input_date.month == next_28_days.month:
        next_28_days = next_28_days + timedelta(days=5)
    last_day_of_month = next_28_days.replace(day=1) - timedelta(days=1)
    first_day_of_next_month = last_day_of_month + timedelta(days=1)
    return last_day_of_month, first_day_of_next_month


def export_excel(filtered_data, column_names, column_field_names, file_name):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{file_name}.xlsx"'

    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Write header row (if needed)
    ws.append(column_names)

    # Write data rows
    for item in filtered_data:
        row = []
        for col_name in column_field_names:
            val = getattr(item, col_name)
            row.append(val)
        ws.append(row)

        # Save the workbook to the response
    wb.save(response)

    return response


def export_csv(filtered_data, column_names, column_field_names, file_name):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{file_name}.csv"'

    # Create a CSV writer
    writer = csv.writer(response)

    # Write header row (if needed)
    writer.writerow(column_names)

    # Write data rows
    for item in filtered_data:
        row = []
        for col_name in column_field_names:
            val = getattr(item, col_name)
            row.append(val)
        writer.writerow(row)

    return response


def get_formatted_date(input_date):
    if input_date:
        return "{0}-{1}-{2}".format(input_date.year, input_date.month, input_date.day)
    return None


def get_process(cur_proc):
    return ",".join(proc.name for proc in cur_proc.all()) if cur_proc.count() > 0 else None


def read_file_format_output(file, skip_cols=None, map_columns=None, return_wb_ws=False):
    wb = load_workbook(file)
    ws = wb.active

    xl_dict = {}
    for col in ws.iter_cols(values_only=True):
        col_name = str(col[0]).strip()
        if skip_cols:
            if col_name in skip_cols:
                continue
        if map_columns:
            col_name = map_columns.get(col_name, col_name)
        xl_dict[col_name] = list(col[1:])
    if return_wb_ws:
        return xl_dict, wb, ws
    return xl_dict


def read_test_question_file_format_output(file):
    wb = load_workbook(file)
    ws = wb.active

    xl_dict = {}
    count = 0
    for row in ws.iter_rows(values_only=True):
        if count == 4:
            break
        xl_dict[str(row[0]).lower()] = row[1]
        if row[3] and row[4]:
            xl_dict[str(row[3])] = row[4]
        count += 1

    for col in ws.iter_cols(values_only=True):
        xl_dict[str(col[4]).lower()] = col[5:]
    return xl_dict


def format_text_answer(ans):
    if ans:
        ans = str(re.sub(r'[^A-Za-z0-9 ]', '', ans)).lower()
        ans = " ".join(i for i in ans.split(" ") if str(i).isalnum())
        return ans
    return None


def get_time_difference(time1, time2):
    total_seconds1 = time1.hour * 3600 + time1.minute * 60 + time1.second
    total_seconds2 = time2.hour * 3600 + time2.minute * 60 + time2.second

    # Calculate the time difference in seconds
    time_difference_seconds = total_seconds2 - total_seconds1

    # Convert the time difference back to hours, minutes, and seconds
    hours = time_difference_seconds // 3600
    minutes = (time_difference_seconds % 3600) // 60
    seconds = time_difference_seconds % 60
    return hours, minutes, seconds


def get_diff_hours_minutes(date1, date2):
    time_difference_seconds = (date2.astimezone(pytz.utc) - date1.astimezone(pytz.utc)).seconds
    days = (date2.astimezone(pytz.utc) - date1.astimezone(pytz.utc)).days
    hours = time_difference_seconds // 3600
    minutes = (time_difference_seconds % 3600) // 60
    seconds = time_difference_seconds % 60
    hours = hours + days * 24
    return hours, minutes, seconds


def format_hour_min_second(h, m, s):
    h = str(h)
    m = str(m)
    s = str(s)
    if len(h) == 1:
        h = f"0{h}"
    if len(m) == 1:
        m = f"0{m}"
    if len(s) == 1:
        s = f"0{s}"
    return h, m, s


def add_hour_minute_to_time(hours, minutes, time1):
    time1 = time.fromisoformat(time1)
    add_time = timedelta(hours=hours, minutes=minutes)
    return (datetime.combine(date.today(), time1) + add_time).time()


def reduce_hour_minute_to_time(hours, minutes, time1):
    time1 = time.fromisoformat(time1)
    add_time = timedelta(hours=hours, minutes=minutes)
    return (datetime.combine(date.today(), time1) - add_time).time()


def is_management(emp):
    # TODO make it non-hard-coded
    if emp.team is None:
        return None
    return emp.designation.category.name in ["VP", "Management"]


def get_hour_min_second_format(timedelta_value):
    h, m, s = str(timedelta_value).split(":")
    return f"{int(h)}:{int(m)}:{int(float(s))}"


def is_hr(emp):
    return emp.designation.department.dept_id == "hr"

def is_cc(emp):
    return emp.designation.department.dept_id=="cc"

def convert_seconds(seconds):
    try:
        seconds = int(seconds)

        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        secs = int(seconds % 60)
        # logger.info(f"here float {seconds}")
        return f"{hours}:{minutes}:{secs}"
    except:
        # logger.info(f"here time {seconds}")
        return seconds


def is_manager_and_above(emp):
    if emp.team is None:
        return None
    return emp.designation.category.name in ["Manager", "VP", "Management"]


def split_week_off_and_return(request):
    week_offs = request.data.get("week_off", "")
    week_offs = week_offs.split(",")
    return week_offs


def get_start_date_end_date_week_offs_start_end_time(CreateAttendanceScheduleSerializer, request):
    start_date = request.data.get("start_date")
    end_date = request.data.get("end_date")
    if start_date is None or end_date is None:
        raise ValueError("Provide proper start date and end date")

    serializer = CreateAttendanceScheduleSerializer(data=request.data)
    if not serializer.is_valid():
        return_error_response(serializer, raise_exception=True)
    start_date = datetime.fromisoformat(start_date).date()
    end_date = datetime.fromisoformat(end_date).date()
    start_time = serializer.validated_data["start_time"]
    end_time = serializer.validated_data["end_time"]
    week_offs = split_week_off_and_return(request)

    return start_date, end_date, week_offs, start_time, end_time


def update_schedule(Attendance, created_by, profile, start_date, end_date, week_offs, start_time, end_time, comment,
                    is_training, Schedule, joining_date=None, att_list=[], att_up_list=[], sch_list=[], upload=False,
                    create_type="GEN"):
    limit_start_date = joining_date if joining_date else datetime.today().date()
    if start_date < limit_start_date or end_date < limit_start_date:
        raise ValueError(
            "start-date {0} and end-date {1} for employee {2} must be greater than today".format(start_date, end_date,
                                                                                                 profile.emp_id))
    if upload:
        att_list = []
        att_up_list = []
        sch_list = []
    dates = get_dates_range(start_date, end_date)
    week_offs = [str(i).strip() for i in week_offs]
    attendance__dates = list(
        Attendance.objects.filter(profile=profile, date__gte=start_date, date__lte=end_date).exclude(
            status__in=["Scheduled", "Week Off"]).values_list("date", flat=True))
    message = "No Changes made"
    week_offs_str = " ".join(str(wk).strip() for wk in week_offs)
    for dt in dates:
        if dt not in attendance__dates:
            if dt.strftime("%A") in week_offs:
                stat = "Week Off"
            else:
                stat = "Scheduled"
            # Updating the sch_upd_by when schedule is updated
            defaults = {"date": dt, "start_time": start_time, "end_time": end_time, "status": stat,
                        "profile": profile, "comment": comment, "is_training": is_training, "sch_upd_by": created_by}
            attendance = Attendance.objects.filter(profile=profile, date=dt).last()
            if not attendance:
                att_list.append(Attendance(**defaults))
            message = "Schedule create is successful"
            if attendance:
                if comment is None:
                    raise ValueError(
                        "Please provide comment, Schedule exists for {0} on {1}, please choose updating the schedule".format(
                            profile.emp_id, str(dt)))
                for key, value in defaults.items():
                    setattr(attendance, key, value)
                setattr(attendance, 'updated_at', datetime.now())
                att_up_list.append(attendance)
                message = "Schedule update is successful"
    sch_list.append(
        Schedule(profile=profile, start_date=start_date, end_date=end_date, start_time=start_time,
                 end_time=end_time, week_off=week_offs_str, comment=message,
                 created_by=created_by, is_training=is_training, create_type=create_type, emp_id=profile.emp_id,
                 rm1_id=profile.team.manager.emp_id, created_by_emp_id=created_by.emp_id))
    if upload:
        message = create_att_list_upd_list_sch_list(Attendance, Schedule, att_list, att_up_list, sch_list)
        return [], [], []
    return att_list, att_up_list, sch_list


def create_att_list_upd_list_sch_list(Attendance, Schedule, att_list, att_up_list, sch_list):
    message = "no changes made"
    actions = []
    if len(att_list) > 0:
        actions.append("create")
        Attendance.objects.bulk_create(att_list)
        logger.info("create is successful")
    if len(att_up_list) > 0:
        Attendance.objects.bulk_update(att_up_list,
                                       ["date", "start_time", "end_time", "status", "profile", "comment", "is_training",
                                        "sch_upd_by", "updated_at"])
        logger.info("update is successful")
        actions.append("update")
    if len(sch_list) > 0:
        Schedule.objects.bulk_create(sch_list)
        logger.info("schedule create is successful")
    if len(actions) > 0:
        message = "Attendance " + " ".join(i for i in actions) + " is successful"
    return message


def check_if_candidate_rejected(candidate):
    if candidate.status == "Rejected":
        raise ValueError("Candidate application has been rejected please reapply for new position")


def get_candidate_interview_timing(candidate, application, test):
    from interview_test.models import CandidateInterviewTestTiming
    cd_test_timing = CandidateInterviewTestTiming.objects.filter(candidate=candidate, application=application,
                                                                 erf=application.position_applied).last()
    return cd_test_timing


class Gpt():

    def __init__(self, model="gpt-4-turbo-preview"):
        self.client = OpenAI(api_key=settings.OPENAI_API_KEY)
        self.model = model

    def make_gpt_call(self, content_condition, search_content):
        response = self.client.chat.completions.create(
            model=self.model,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system",
                 "content": content_condition
                 },
                {"role": "user", "content": str(search_content)}
            ]
        )
        return response.choices[0].message.content


def get_gpt():
    global gpt
    if gpt is None:
        gpt = Gpt()
    return gpt


def validate_answers(answer, question):
    max_score = 10
    # Prepare the prompt
    condition = f"Evaluate the following answer to the question and give a score out of {max_score}, return the result as {{score: number}} in JSON."

    search_content = f"[Question]\n{question}\n"
    search_content += f"[Answer]\n{answer}\n"

    chatgpt_score = None

    while chatgpt_score is None:
        try:
            response = get_gpt().make_gpt_call(condition, search_content)
            chatgpt_score = json.loads(response)

            if "score" not in chatgpt_score or not isinstance(chatgpt_score["score"], (int, float)):
                chatgpt_score = None
                logger.info(f"Invalid score response, retrying...")
                time.sleep(1)

        except (json.JSONDecodeError, KeyError, ValueError) as e:
            logger.error(f"Error: {e}. Retrying...")
            chatgpt_score = None
            time.sleep(1)

    return chatgpt_score["score"]


def evaluate_essay_by_chat_gpt_in_json_format(text):
    try:
        if text is None or str(text) == "":
            return {"feedback": "There is no input text"}
        response = get_gpt().make_gpt_call(
            "Following is an essay text, give the text feedback considering the grammar and comprehension, feedback within 200 characters, output should be in json_format with feedback key",
            text)
        if response[0] != "{":
            return {"feedback": response}
        return json.loads(response)
    except Exception as e:
        logger.info("failed to evaluate_essay_by_chat_gpt_in_json_format, contact admin {0}".format(str(e)))
        return {"feedback": "internal-error please contact admin"}


def calculate_remaining_duration(actual_test_end_time):
    current_time = tz.localize(datetime.now().replace(microsecond=0))
    if actual_test_end_time.tzinfo is None:
        actual_test_end_time = tz.localize(actual_test_end_time)
    value = actual_test_end_time - current_time
    if value.days < 0:
        return 0
    return value.seconds


def create_candidate_test_section(cand_qa_sec_list, candidate, application, test):
    from interview_test.views import get_candidate_qa_sec_list
    from interview_test.models import CandidateQASection, TestQA

    if cand_qa_sec_list is None:
        cand_qa_sec_list = get_candidate_qa_sec_list(candidate, application, test)
    if len(cand_qa_sec_list) == 0:
        from interview_test.models import TestSection
        test_sections = TestSection.objects.filter(test=test)
        count = 0
        for t_section in test_sections:
            section_status = "Pending"
            if (count == 0 and not test.has_essay) or t_section.is_essay:
                section_status = "Current"
            total_testqa = TestQA.objects.filter(test_section=t_section)
            stats = total_testqa.aggregate(max_score=Sum("qa__max_score"))
            max_score = stats.get("max_score", 0)
            cand_qa_sec = CandidateQASection.objects.create(
                total_qns=total_testqa.count(), candidate=candidate, application=application, test_section=t_section,
                section_status=section_status, max_score=max_score, erf=application.position_applied)
            count += 1


def get_teams_under_me(manager, teams_list):
    from team.models import Team
    # print(manager, manager.team.manager)
    from mapping.models import Profile
    # teams = list(
    #     Profile.objects.filter(team__manager=manager).values_list("team", flat=True).distinct())
    teams = list(set(Team.objects.filter(manager=manager, is_active=True).values_list("id", flat=True).distinct()))
    # print("teams", teams)
    if len(teams) == 0:
        return teams_list
    teams_list = teams_list + teams

    for team_id in teams:
        team = Team.objects.get(id=team_id)
        profiles = Profile.objects.filter(team=team)
        for profile in profiles:
            if profile != manager:
                teams_list = get_teams_under_me(profile, teams_list)

    return teams_list


def update_my_team(managers=None):
    from team.models import Team
    from mapping.models import Profile
    teams = Team.objects.all()
    agents = Profile.objects.filter(~Q(id__in=list(teams.values_list("manager__id", flat=True))))
    for agent in agents:
        agent.my_team.clear()

    if managers is None:
        teams = teams.filter(is_active=True)
    else:
        teams = teams.filter(manager__id__in=managers, is_active=True)

    for team in teams:
        if team.manager:
            # logger.info("before {0}".format(team.manager))
            my_teams = get_teams_under_me(team.manager, [])
            # logger.info("updating {0}".format(team.manager))
            existing_my_team = team.manager.my_team.all()

            for mteam in existing_my_team:
                if mteam not in my_teams:
                    team.manager.my_team.remove(mteam)
            for mteam in my_teams:
                team.manager.my_team.add(mteam)


def get_iexact_for_list(field_name, params=[]):
    q = Q()
    for param in params:
        tmp = {f"{field_name}__iexact": param}
        q |= Q(**tmp)
    return q


def validate_current_century(value):
    val = datetime.today().date() - value
    year = int(val.days / 365)
    if year >= 100 or year < 18:
        raise ValidationError('Candidate must be greater than 18 years and less than 100 years')


def update_leave_balance(profile, transaction, no_of_days, leave_type="PL"):
    from ams.models import LeaveBalance
    leave_bal = LeaveBalance.objects.filter(profile=profile).last()
    if not leave_bal:
        raise ValueError(
            "Leave Balance for {} doesn't exist. Contact Admin.".format(profile.emp_id))
    msg = f"credited back {no_of_days} {leave_type} leaves for {transaction}"
    if leave_type == "PL":
        leave_bal.paid_leaves += no_of_days
    else:
        leave_bal.sick_leaves += no_of_days
    leave_bal.total += no_of_days
    leave_bal.save()
    from ams.models import LeaveBalanceHistory
    LeaveBalanceHistory.objects.create(sl_balance=leave_bal.sick_leaves, pl_balance=leave_bal.paid_leaves,
                                       profile=profile, leave_type=leave_type, transaction=msg, no_of_days=no_of_days,
                                       total=leave_bal.total)


def create_mapped_teams(main_teams, from_employee, to_employee, msg, created_by):
    from mapping.models import Profile
    my_teams = from_employee.my_team.all()
    update_qms = False
    if settings.ENABLE_QMS_INTEGRATION and check_server_status(settings.QMS_URL):
        update_qms = True

    if my_teams.count() > 0:
        from_employee.my_team.clear()
    if main_teams is None:
        from team.models import Team
        main_teams = Team.objects.filter(manager=from_employee)
    if main_teams.count() > 0:
        from mapping.models import MappedTeams
        map_teams = MappedTeams.objects.create(from_employee=from_employee, reason=msg, to_employee=to_employee,
                                               created_by=created_by)

        for tm in main_teams:
            map_teams.mapped_teams.add(tm)
            tm.manager = to_employee
            tm.save()
            # Updating RM1, RM2, RM3 of employees under the person Mapped (in QMS)
            if update_qms:
                emp_in_team = Profile.objects.filter(team=tm)
                for emp in emp_in_team:
                    update_profile_in_qms(emp.emp_id)
        return map_teams
    return None


def get_all_connected_managers(employee, mgrs_list=[]):
    mgr = employee.team.manager if employee.team.manager and employee.team else None
    if mgr is None or mgr == employee:
        if mgr:
            mgrs_list.append(mgr.id)
        return mgrs_list
    mgrs_list.append(mgr.id)
    return get_all_connected_managers(mgr, mgrs_list)


def get_connected_managers_list(employees):
    mgrs_list = []
    for employee in employees:
        mgrs_list.append(employee.id)
        mgrs_list = get_all_connected_managers(employee, mgrs_list)
    return list(set(mgrs_list))


def get_selected_emp_under_team(usr, team, emp_ids, consider_active=True):
    team_ids = get_team_ids(usr)
    if team_ids is None:
        raise ValueError("there are no employees under your team")
    from mapping.models import Profile
    if team is None:
        raise ValueError("Please provide the team information")
    elif team == "__all_":
        # Excluding Inactive profiles from Creating/Updating Schedule
        profiles = Profile.objects.filter(team__id__in=team_ids)
    else:
        team = str(team).split(",")
        if emp_ids is None:
            raise ValueError("Please provide the employees list")
        elif "__all_" in emp_ids:
            profiles = Profile.objects.filter(team__id__in=team)
        else:
            emp_ids = emp_ids.split(",")
            profiles = Profile.objects.filter(emp_id__in=emp_ids)
    if consider_active:
        return profiles.filter(is_active=True)
    return profiles


def create_profile_in_qms(emp_id):
    try:
        if settings.ENABLE_QMS_INTEGRATION:
            url = settings.QMS_URL + '/ProfileView/' + str(emp_id)
            from mapping.models import Profile
            profile = Profile.objects.get(emp_id=emp_id)
            if not profile:
                raise ValueError("Profile with this emp_id not found.")
            from mapping.serializer import QMSProfileSerializer
            serializer = QMSProfileSerializer(profile)
            try:
                qms_profile = requests.post(url, json=serializer.data)
            except Exception as e:
                raise ValueError("Issue faced while creating Profile in QMS")
            logger.info('Profile created in QMS')
    except Exception as e:
        logger.info("exception at create_profile_in_qms {0} ".format(traceback.format_exc(5)))


def update_profile_in_qms(emp_id):
    try:
        if settings.ENABLE_QMS_INTEGRATION:
            url = settings.QMS_URL + '/ProfileView/' + str(emp_id)
            from mapping.models import Profile
            profile = Profile.objects.get(emp_id=emp_id)
            if not profile:
                raise ValueError("Profile with this emp_id not found.")
            from mapping.serializer import QMSProfileSerializer
            serializer = QMSProfileSerializer(profile)
            try:
                qms_profile = requests.post(url, json=serializer.data)
            except Exception as e:
                raise ValueError("Issue faced while updating Profile in QMS")
            logger.info('Profile updated in QMS')
    except Exception as e:
        logger.info("exception at update_profile_in_qms {0} ".format(traceback.format_exc(5)))


def document_file(instance, filename):
    today = datetime.today()
    uuid_value = uuid.uuid4()
    if settings.IS_PUBLIC_SERVER:
        try:
            sub_directory = f"documents/{today.year}/{today.month}/{today.day}"
            # Save the file in the database server (192.168.10.1)
            # Replace '/root/hrms/HRMS-BOBO-BE-BOBO/media/' with the actual path on the server
            new_filename = f"{sub_directory}/{uuid_value}/{filename}"
            response = requests.post(settings.SERVER_PUBLIC_FILE_SAVE_ENDPOINT, json={
                "path": new_filename,
                "file": base64.b64encode(instance.file.read()).decode('utf-8')  # Convert bytes to string
            })
        except Exception as e:
            logger.info("exception at interview documents_file while saving the file in the server {0}".format(
                traceback.format_exc(5)))

    return '/'.join(
        ["documents", str(today.year), str(today.month), str(today.day), str(uuid_value), filename])


def get_careers_email_message(subject, email_template, to=[], reply_to=[]):
    from_email, email_connection = get_email_connection('careers')
    return EmailMessage(subject, email_template, from_email, to, connection=email_connection)


def get_erf_email_message(subject, email_template, to=[], reply_to=[]):
    from_email, email_connection = get_email_connection('erf')
    return EmailMessage(subject, email_template, from_email, to, connection=email_connection)


def get_development_email_message(subject, email_template, to=[], reply_to=[]):
    from_email, email_connection = get_email_connection('development')
    return EmailMessage(subject, email_template, from_email, to, connection=email_connection)


def send_email_message(email_msg):
    email_msg.content_subtype = 'html'
    try:
        if settings.SEND_EMAIL:
            email_msg.send()
    except SMTPAuthenticationError as e:
        logger.critical("Exception occurred at send_email_message {0}".format(e))
    except Exception as e:
        logger.critical("Exception occurred at send_email_message {0}".format(e))


def get_default_email_connection():
    return settings.EMAIL_HOST_USER, get_connection(
        host=settings.EMAIL_HOST,
        port=settings.EMAIL_PORT,
        username=settings.EMAIL_HOST_USER,
        password=settings.EMAIL_HOST_PASSWORD,
        use_tls=settings.EMAIL_USE_TLS
    )


def get_email_connection(from_email_name):
    email_accounts = settings.EMAIL_ACCOUNTS

    if not email_accounts:
        return get_default_email_connection()

    else:
        if from_email_name in email_accounts:
            account = email_accounts[from_email_name]
            email_connection = get_connection(
                host=settings.EMAIL_HOST,
                port=settings.EMAIL_PORT,
                username=account['host_user'],
                password=account['host_password'],
                use_tls=settings.EMAIL_USE_TLS
            )
            return account["host_user"], email_connection

        else:
            return get_default_email_connection()


def update_email(email, candidate):
    # TODO replace with check_for_integrity_error
    try:
        if email:
            setattr(candidate, 'email', email.lower())
            candidate.save()
    except IntegrityError as e:
        raise ValueError(
            "Candidate with this {0} already exists".format(re.split("\,| |\.", str(e))[-1].replace("_", " ")))


def check_for_integrity_error(field, field_value, Obj):
    if field_value:
        try:
            setattr(Obj, field, field_value)
            Obj.save()
            return Obj
        except IntegrityError as e:
            raise ValueError(
                "{0} with this {1} already exists".format(Obj.__class__.__name__, str(field).replace("_", " ")))


def validate_mobile_no(value):
    validate_number_with_fixed_size(value, 'mobile number', 10)


def validate_alphanumeric(value):
    value = value.split(" ")
    for i in value:
        if not str(i).isalnum():
            raise ValidationError("Account name cannot have special characters")


def validate_bnk_account_no(value):
    if not str(value).isnumeric():
        raise ValidationError("Bank account number should be numeric")


def validate_ifsc(value):
    if len(str(value)) != 11:
        raise ValidationError("IFSC Code must be of length 11")
    if not str(value).isalnum() or not str(value)[:4].isalpha():
        raise ValidationError("Please provide valid IFSC Code")


def validate_model_field(value, func_name):
    try:
        func_name(value)
    except ValidationError as e:
        error_message = e.detail[0] if isinstance(e.detail, list) else e.detail
        raise ValueError(error_message)
    except DjangoValidationError as e:
        raise ValueError(str(e.message))


def get_manager(emp, mgrs):
    if not (emp.team and emp.team.manager) or (
            emp.team.manager and emp.team.manager.team.manager and emp == emp.team.manager.team.manager):
        if emp == emp.team.manager:
            mgrs.append(emp.id)
        return mgrs
    mgrs.append(emp.team.manager.id)
    return get_manager(emp.team.manager, mgrs)


def get_all_managers(emp):
    return list(set(get_manager(emp, [])))


def get_all_managers_include_emp(emp_list):
    mgrs_list = []
    for emp in emp_list:
        mgrs_list.append(emp.id)
        mgrs_list = get_manager(emp, mgrs_list)
    return list(set(mgrs_list))


def check_if_date_gte_25():
    # Checks if today is within the last 5 days of the month
    last_day_of_cur_month, _ = get_last_day_of_cur_month_and_first_day_of_next_month()
    return datetime.today().date() >= (last_day_of_cur_month - timedelta(days=4)) or datetime.today().day == 1
    # return False


def get_25th_date():
    today = datetime.today().date()
    if today.day == 1:
        return (today.replace(day=1) - timedelta(days=1)).replace(day=25)
    return today.replace(day=25)


def get_1st_date():
    # very specific to attendance
    today = datetime.today().date()
    if today.day == 1:
        return (today - timedelta(days=1)).replace(day=1)
    return today.replace(day=1)


def validate_team_and_rm1_rm2_rm3(usr, profile, my_team, rm3=True):
    if profile.team.id not in my_team:
        raise ValueError("emp-id {0} with team {1} doesn't belong to your team".format(
            profile.emp_id, profile.team.name))

    mgr = profile.team.manager
    if mgr is None or mgr.team.manager is None or mgr.team.manager.team.manager is None:
        raise ValueError(
            "RM1 RM2 RM3 is not assigned properly for the employee {0}".format(profile.emp_id))
    if rm3:
        if not (usr == mgr or usr == mgr.team.manager or usr == mgr.team.manager.team.manager):
            raise ValueError("This activity can be done only the employees respective RM1, RM2 or RM3")
    else:
        if not (usr == mgr or usr == mgr.team.manager):
            raise ValueError("This activity can be done by only the employees respective RM1 and RM2")


def validate_aadhaar_no(value):
    validate_number_with_fixed_size(value, 'aadhaar number', 12)


def check_if_date_is_excepted(date1):
    if check_if_att_exc_allowed():
        if date1 in get_excluded_dates():
            return True
    return False


def get_excluded_dates():
    if not check_if_att_exc_allowed():
        return []
    return get_dates_range(settings.ATT_EXC_START_DATE,
                           settings.ATT_EXC_END_DATE)


def check_if_att_exc_allowed():
    today = datetime.today().date()
    if settings.ATT_EXC_ALLOWED and today == settings.ATT_EXC_DATE:
        return True
    return False


def get_column_names_only(model):
    model_fields = model._meta.get_fields()
    column_names = [field.name for field in model_fields if field.concrete and not isinstance(field, (
        models.FileField, models.ManyToManyField, models.ForeignKey, RichTextField))]
    logger.info(f"get_column_names_only {column_names}")
    return column_names


def get_column_names(model):
    model_fields = model._meta.get_fields()
    column_names = [field.name for field in model_fields if field.concrete and not isinstance(field, (
        models.FileField, models.ManyToManyField, models.ForeignKey, RichTextField))]
    related_cols = [field.name for field in model_fields if field.concrete and isinstance(field, (
        models.ManyToManyField, models.ForeignKey))]

    return column_names, related_cols


def decrypt_obj(model, data, obj):
    model_fields = model._meta.get_fields()
    for field in model_fields:
        if field.concrete and isinstance(field, EncryptedCField):
            data[field.name] = get_decrypted_value(getattr(obj, field.name))
    return data


def get_column_names_with_foreign_keys_separate(model):
    model_fields = model._meta.get_fields()
    column_names = [field.name for field in model_fields if field.concrete and not isinstance(field, (
        models.FileField, models.ManyToManyField, models.ForeignKey, RichTextField))]
    foreign_keys = [field.name for field in model_fields if field.concrete and isinstance(field, models.ForeignKey)]
    related_cols = [field.name for field in model_fields if field.concrete and isinstance(field, (
        models.ManyToManyField, models.ForeignKey))]

    return column_names, foreign_keys, related_cols


def parse_boolean_value(value):
    if type(value) is bool:
        return value
    return json.loads(value) if value and value in ["true", "false", "1", "0"] else None


def get_my_team_process(profile):
    processes = {}
    for i in profile.my_team.all():
        tm_mgr = [i.id, i.name + " ( " + i.manager.emp_id + " - " + i.manager.full_name + " )"]
        if i.base_team.name in processes:
            processes[i.base_team.name].append(tm_mgr)
        else:
            processes[i.base_team.name] = [tm_mgr]
    return processes


def get_my_team(profile):
    if profile.designation.category.name in ["VP", "Management List"]:
        from team.models import Team
        my_team = list(Team.objects.all().values_list("id", "name", named=True))
    else:
        my_team = [[i.id, i.name] for i in
                   profile.my_team.all()] if profile.my_team and profile.my_team.all().count() > 0 else None
    return my_team


def check_for_60th_day(end_date):
    next_60th_day = datetime.today().date() + timedelta(days=60)
    if isinstance(end_date, datetime):
        end_date = end_date.date()
    if end_date > next_60th_day:
        raise ValueError(
            "Creation/update of schedule is allowed for only next 60 days, considering today the last 60th day is {0}".format(
                next_60th_day))


def get_month_number(month_name):
    map_month_names = {"january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6, "july": 7, "august": 8,
                       "september": 9, "october": 10, "november": 11, "december": 12}
    month_number = map_month_names.get(month_name.lower())
    if month_number is None:
        raise ValueError("Invalid month name {0}".strip(month_name))
    return month_number


def get_date_value(date_val):
    if isinstance(date_val, datetime):
        return date_val.date()
    elif isinstance(date_val, str):
        return datetime.fromisoformat(date_val)
    else:
        raise ValueError("invalid date_val for get_date_value")


def get_date_time_value(date_val):
    if isinstance(date_val, datetime):
        return date_val
    elif isinstance(date_val, str):
        return datetime.fromisoformat(date_val)
    else:
        raise ValueError("invalid date_val for get_date_time_value")


def check_if_user_is_inactive(usr):
    if not usr.is_active:
        raise ValueError("Approve operation is not permitted the user is in {0} status".format(usr.status))


def valid_email_and_aadhaar(email, aadhaar_no, candidate):
    validate_model_field(email, validate_email)
    candidate = check_for_integrity_error("email", email.lower(), candidate)
    validate_model_field(aadhaar_no, validate_aadhaar_no)
    candidate = check_for_integrity_error("aadhaar_no", aadhaar_no, candidate)
    return candidate


def get_formatted_choice(choice_list):
    return " / ".join(i[0] for i in choice_list)


def get_attrition_profiles_from_1st_last_month():
    date1 = (datetime.today().date().replace(day=1) - timedelta(days=1)).replace(day=1)
    from mapping.models import Profile
    return Profile.objects.filter(status__in=["Attrition"], last_working_day__gte=date1)


def non_attrition_profiles():
    from mapping.models import Profile
    return Profile.objects.exclude(status__in=["Attrition"]).exclude(is_superuser=True)


def update_qms_profiles():
    if settings.ENABLE_QMS_INTEGRATION and check_server_status(settings.QMS_URL):
        from mapping.models import Profile
        profiles = Profile.objects.all()
        for profile in profiles:
            # create_profile_in_qms(profile.emp_id)
            update_profile_in_qms(profile.emp_id)


def replace_with_underscore(obj_data, obj, field):
    if isinstance(obj, dict):
        value = obj[field]
    else:
        value = getattr(obj, field)
    obj_data[field.replace("_", " ").capitalize()] = value if value else "-"
    return obj_data


def format_dict_data(obj_data, obj, fields_list):
    for field in fields_list:
        if isinstance(obj, dict):
            obj_data = replace_with_underscore(obj_data, obj, field)

        elif isinstance(obj, list):
            for iobj in obj:
                obj_data = replace_with_underscore(obj_data, iobj, field)
    return obj_data


def format_obj_data(obj_data, obj, fields_list):
    for field in fields_list:
        obj_data = replace_with_underscore(obj_data, obj, field)
    return obj_data


def send_custom_email(email_data_list, subject, template_name):
    try:

        # Create a dictionary with the values to be passed to the template

        # Render the email content from the template with the dictionary
        email_content = render_to_string(template_name, {
            'email_data_list': email_data_list  # [(header_text, email_data),...]
        })

        # Send the email
        send_mail(
            subject,
            email_content,
            settings.DEFAULT_FROM_EMAIL,
            ["lohit.n@expertcallers.com"],
            fail_silently=False,
            html_message=email_content,  # Enable HTML content in the email
        )
    except Exception as e:
        logger.info("exception at send_custom_email {0}".format(traceback.format_exc(5)))


def check_server_status(server_ip):
    try:
        server_ip = server_ip.replace("http://", "").replace("https://", "")
        if len(server_ip) < 3:
            return False
        if str(server_ip)[0].isnumeric():
            server_ip = server_ip.split(":")[0]
        # Run the ping command with a timeout of 5 seconds
        subprocess.run(['ping', '-c', '1', '-W', '1', server_ip], check=True)
        return True
    except subprocess.CalledProcessError as e:
        logger.info(f"error at subprocess call -{e}")
        return False


def get_limited_breaks():
    from mapping.models import MiscellaneousMiniFields
    break_type_limited = list(
        MiscellaneousMiniFields.objects.filter(field='break_type_limited').values_list('content', flat=True))
    return [] if len(break_type_limited) == 0 else [i.strip() for i in str(break_type_limited[0]).split(",")]


# Replace 'example.com' with the IP address or hostname of the server you want to ping

def get_ip_address(request):
    user_ip = request.META.get('HTTP_X_REAL_IP', None)
    logger.info(
        f"real,forwarded{request.META.get('HTTP_X_REAL_IP', None)}, {request.META.get('HTTP_X_FORWARDED_FOR', None)}")

    # You can also handle cases where the IP address is not available in REMOTE_ADDR
    if not user_ip:
        # Attempt to get the IP address from other headers or methods
        user_ip = request.META.get('HTTP_X_FORWARDED_FOR', None)
    return user_ip


def get_custom_emp_response(instance):
    return [instance.emp_id,
            instance.full_name or "" + " - " + str(instance.emp_id) + " - " + str(instance.designation.name or "")]


def check_appraisal_last_date():
    from mapping.models import MiscellaneousMiniFields
    last_date = MiscellaneousMiniFields.objects.filter(field="appraisal_last_date").first()
    last_date = datetime.fromisoformat(
        last_date.content).date() if last_date and last_date.content else datetime.now().replace(
        day=15, month=4).date()
    if datetime.now().date() > last_date:
        raise ValueError(f"Appraisal for {datetime.now().year} is closed on {str(last_date)}")


def get_appraisal_eligible_date():
    from mapping.models import MiscellaneousMiniFields
    app_date = MiscellaneousMiniFields.objects.filter(field="appraisal_eligible_date").last()
    if app_date is None:
        raise ValueError("appraisal eligible date is not added")
    return datetime.fromisoformat(app_date.content).date()


def get_exclude_appraisal_emps():
    from appraisal.models import ExcludeList
    exclude_list = ExcludeList.objects.all()
    if exclude_list.count() == 0:
        return []
    profiles = []
    app_date = get_appraisal_eligible_date()
    for ex in exclude_list:
        from mapping.models import Profile
        profiles += list(
            Profile.objects.filter(status="Active", date_of_joining__lte=app_date, team__base_team=ex.process,
                                   designation=ex.designation).values_list("emp_id", flat=True))
    return profiles


def verify_candidate_email_otp_info(request):
    otp = request.data.get('email_otp') or None
    email = request.data.get('email') or None

    if email is None or otp is None:
        raise ValueError("Email and otp is Required")
    email = email.lower()
    from interview.models import CandidateOTPTest
    cand_otp_obj = CandidateOTPTest.objects.filter(verify_otp=otp, email=email).first()
    if cand_otp_obj is None:
        raise ValueError("Invalid otp/email data, request for otp verification again")
    ninty_minutes_earlier = datetime.now() - timedelta(minutes=90)
    if tz.localize(ninty_minutes_earlier) > cand_otp_obj.updated_at:
        raise ValueError("The OTP is expired, please request for new one")
    return otp, email, cand_otp_obj


get_all_profile_serializer_map_columns = {"gender": "onboard__gender", "base_team": "team__base_team__name",
                                          "team": "team__name",
                                          "designation": "designation__name",
                                          "department": "designation__department__name",
                                          "rm1_full_name": "team__manager__full_name",
                                          "rm1_id": "team__manager__emp_id",
                                          "rm2_id": "team__manager__team__manager__emp_id",
                                          "rm3_id": "team__manager__team__manager__team_manager__emp_id",
                                          "dob_month": "dob__month", "aadhaar_no": "onboard__aadhaar_no",
                                          "pan_no": "onboard__pan_no", "contact_no": "onboard__candidate__mobile",
                                          "rm2_full_name": "team__manager__team__manager__full_name",
                                          "rm3_full_name": "team__manager__team__manager__team__manager__full_name",
                                          "erf_id": "onboard__appointment_letter__application__position_applied__id"}


def calc_esi(gross):
    esi = round(gross * (0.75 / 100), 0) if gross <= settings.ESI_LIMIT else 0
    esi_emr = round(gross * (3.25 / 100), 0) if gross <= settings.ESI_LIMIT else 0
    return esi, esi_emr


def calc_pf(basic, pay_grade_calc):
    pf_emp = round(basic * (pay_grade_calc.pf_emp_perc / 100), 0) \
        if round(basic, 0) <= settings.PF_GROSS_LIMIT else settings.PF_FIXED_DEDUCTION_AMOUNT
    pf_emr = round(basic * (pay_grade_calc.pf_emr_perc / 100), 0) \
        if round(basic, 0) <= settings.PF_GROSS_LIMIT else settings.PF_FIXED_DEDUCTION_AMOUNT
    return pf_emp, pf_emr


def decrypt_salary_fields(data, salary_cols, salary):
    if salary:
        for key in salary_cols:
            if key not in ["pay_grade", "last_update_reason", "opting_for_pf", "created_at", "updated_at"]:
                data[key] = get_decrypted_value(getattr(salary, key))
            else:
                data[key] = str(getattr(salary, key))
            if key == "opting_for_pf":
                data[key] = True if data[key] == "True" else False

    else:
        for key in salary_cols:
            data[key] = None

    return data


def update_salary_history_data(sal_obj, serializer_val_data, from_data, to_data, fields_updated, usr):
    updated = False
    from payroll.models import Salary
    column_names, _ = get_column_names(Salary)
    for key, value in serializer_val_data.items():
        if key not in column_names:
            continue
        from_field_data = getattr(sal_obj, key)
        if from_field_data and key not in ["opting_for_pf", "last_update_reason", "pay_grade"]:
            from_field_data = get_decrypted_value(from_field_data)
            if key not in ["uan", "esic_no", "pf_no", "bank_name",
                           "ifsc_code", "account_number", "name_as_per_bank_acc"]:
                from_field_data = float(from_field_data) if str(from_field_data).isnumeric() else from_field_data
                value = float(value) if value and str(value).isnumeric() else value
            else:
                value = str(value) if value else value

        if not value == from_field_data:
            if (from_field_data is None and value != "") or (value == "" and from_field_data is None):
                continue
            from_data[key] = from_field_data
            to_data[key] = value
            fields_updated.append(key)
        updated = True
        setattr(sal_obj, key, value)

    if updated:
        setattr(sal_obj, 'updated_by', usr)
    sal_obj.save()
    return from_data, to_data, fields_updated


def update_salary_history(salary, request, serializer_val_data, usr, comment):
    from_data = {}
    to_data = {}
    fields_updated = []
    if salary:
        account_number = request.data.get("account_number")
        serializer_val_data["account_number"] = account_number
        from_data, to_data, fields_updated = update_salary_history_data(salary, serializer_val_data, from_data,
                                                                        to_data, fields_updated, usr)
    salary = handle_onboard_salary_account_number(request.data.get("account_number"), salary)
    if len(fields_updated) > 0:
        from payroll.models import UpdateSalaryHistory
        return UpdateSalaryHistory(profile=salary.profile,
                                   candidate=salary.candidate,
                                   field_updated=", ".join(str(i).strip() for i in fields_updated),
                                   updated_by=usr, to_data=to_data, from_data=from_data,
                                   comment=comment if comment else None)
    return None


def handle_onboard_salary_account_number(account_number, salary):
    if account_number == "":
        account_number = None
    try:
        if salary is None:
            raise ValueError("Salary Information does not exist - contact hr")
        setattr(salary, 'account_number', account_number)
        return salary
        # salary.save()
    except IntegrityError as e:
        raise ValueError("Candidate with this salary account number already exists")


def format_partial_salary_information(data, salary, profile):
    cols = ["bank_name", "name_as_per_bank_acc", "account_number", "ifsc_code"]
    for col in cols:
        if salary:
            data[col] = get_decrypted_value(getattr(salary, col)) if salary else None

    data["pay_grade"] = salary.pay_grade if salary else None
    data["emp_id"] = profile.emp_id if profile else None
    if profile:
        data["emp_name"] = profile.full_name
    else:
        data["emp_name"] = salary.candidate.full_name if salary and salary.candidate else None
    data["uan"] = salary.uan if salary else None
    data["pf_no"] = salary.pf_no if salary else None
    data["esic_no"] = salary.esic_no if salary else None
    return data


def get_group_concat_query_part(field_name):
    if settings.DB_VENDOR == "postgresql":
        return f"SELECT STRING_AGG({field_name}, ' | ')"
    elif settings.DB_VENDOR == "mysql":
        return f"SELECT GROUP_CONCAT({field_name} SEPARATOR ' | ')"
    elif settings.DB_VENDOR == "sqlite":
        return f"SELECT GROUP_CONCAT({field_name}, ' | ')"
    else:
        raise ValueError("Unsupported database engine")


def create_notification(title, message=None, link=None, team=[], campaign=[]):
    from article.models import Notifications
    notification_obj = Notifications.objects.create(title=title, message=message, link=link)
    notification_obj.team.set(team)
    notification_obj.campaign.set(campaign)
    return notification_obj


def assigning_notification(notification_obj, added_by=None):
    from mapping.models import Profile
    from article.models import UserNotification
    try:
        profiles = Profile.objects.filter(status="Active").filter(
            Q(id=added_by.id) | Q(team__in=notification_obj.team.all()) | Q(
                designation__category__name__in=["Manager", "VP", "Management"])).distinct()

        UserNotification.objects.bulk_create([
            UserNotification(notification=notification_obj, profile=profile)
            for profile in profiles
        ])
        logger.info("User Notification created successfully")
        return True
    except Exception as e:
        logger.info(f"exception at assigning_notification {traceback.format_exc(5)}")
        return False


def authenticate_article_token(request, article_id=None):
    from article.models import ArticleVersion
    article_token = request.headers.get("X-Article-Token")
    kwarg_article_id = request.parser_context.get("kwargs", {}).get("article_id")
    if kwarg_article_id is None:
        if article_id is None:
            raise ParseError("Process update id is required")
    else:
        article_id = kwarg_article_id

    article = ArticleVersion.objects.filter(id=article_id).first()
    logger.info(f"article token {article_token}")
    request.article = article
    if not article:
        raise ParseError("Invalid Process update id.")
    if not article.article.password:
        return
    if not article_token:
        raise ParseError("Process update Token is required")

    if article.article.article_token != article_token:
        raise ParseError("Invalid Process update token")

    if article.article.token_expires_at and article.article.token_expires_at < timezone.now():
        raise ParseError("Process update Token has expired.")


def article_validate_token(token):
    from article.models import Article
    """
    Validate the token and check its expiration.
    """
    article = Article.objects.filter(article_token=token).first()

    return article


def check_if_method_has_url_access(request, view):
    from mapping.views import HasUrlPermission
    has_url_permission = HasUrlPermission()
    try:
        has_url_permission.has_permission(request, view)
    except:
        raise ValueError("You do not have permissions")


def get_cab_address_obj_by_cab_address_id(cab_address_id):
    from transport.models import CabDetail
    cab_address = CabDetail.objects.filter(id=cab_address_id).first()
    if cab_address is None:
        raise ValueError("The requested cab-address does not exist")
    return cab_address


def update_profile_in_qms3(emp_id):
    try:
        if settings.ENABLE_QMS3_INTEGRATION:
            if settings.is_development:
                sub_url='/user/create_profile/'
            else:
                sub_url='/api/user/create_profile/'
            url = settings.QMS3_URL + sub_url + str(emp_id)
            logger.info(f"QMS3 sending URL {url}")
            from mapping.models import Profile
            profile = Profile.objects.get(emp_id=emp_id)
            if not profile:
                raise ValueError("Profile with this emp_id not found.")
            from mapping.serializer import QMS3ProfileSerializer
            serializer = QMS3ProfileSerializer(profile)
            try:
                qms_profile = requests.post(url, json=serializer.data)
                logger.info(f"data {serializer.data}")
            except Exception as e:
                raise ValueError("Issue faced while updating Profile in QMS")
            logger.info('Profile updated in QMS3')
    except Exception as e:
        logger.info(f"exception at update_profile_in_qms3 {0}".format(traceback.format_exc(5)))
        logger.info(f"exception at update_profile_in_qms3 {e}")


def create_qms3_profiles():
    if settings.ENABLE_QMS3_INTEGRATION and check_server_status(settings.QMS3_URL):
        from mapping.models import Profile
        profiles = Profile.objects.all()
        for profile in profiles:
            emp_id = profile.emp_id
            logger.info(f"before create emp_id {emp_id}")
            update_profile_in_qms3(emp_id)


def create_update_dependencies():
    try:
        if settings.ENABLE_QMS3_INTEGRATION and check_server_status(settings.QMS3_URL):
            from mapping.models import Category, Designation, Department,Profile
            from team.models import Team, Process
            if settings.is_development:
                sub_str='/user/sync_profile_dependencies'
            else:
                sub_str='/api/user/sync_profile_dependencies'
            url = settings.QMS3_URL + sub_str
            from mapping.serializer import CategorySerializer, DepartmentSerializer, DesignationSerializer, \
                ProcessSerializer, TeamSerializer,QMS3ProfileSerializer
            category_serializer = CategorySerializer(Category.objects.all(), many=True)
            department_serializer = DepartmentSerializer(Department.objects.all(), many=True)
            designation_serializer = DesignationSerializer(Designation.objects.all(), many=True)
            process_serializer = ProcessSerializer(Process.objects.all(), many=True)
            team_serializer = TeamSerializer(Team.objects.all(), many=True)
            #---profile sync-----------------
            # profile_serializer=QMS3ProfileSerializer(Profile.objects.filter(is_superuser=False,many=True)
            #-----------------
            json_data = {'category': category_serializer.data, 'department': department_serializer.data,
                         'designation': designation_serializer.data, 'process': process_serializer.data,
                         'team': team_serializer.data
                         # ---profile sync-----------------
                        # ,'profile':profile_serializer.data
                         #--------------
                         }
            try:
                response = requests.post(url, json=json_data)
                logger.info(f"url - sync profile dependencies ---{url}--status - {response.status_code}")
                logger.info("----Completed----------")

            except Exception as e:
                logger.error(f"exception at update profile dependencies {traceback.format_exc(5)}")
                raise ValueError("Issue faced while updating Profile Dependencies in QMS3")
    except Exception as e:
        logger.info(f"exception at update_profile_dependencies_in_qms3 {0}".format(traceback.format_exc(5)))


