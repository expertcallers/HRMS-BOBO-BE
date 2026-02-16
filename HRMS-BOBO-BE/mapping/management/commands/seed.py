# <project>/<app>/management/commands/seed.py
import base64
import csv
import io
import json
import os
import string
import traceback
from datetime import timedelta, datetime, date

import openpyxl
import pyotp
import pytz
import qrcode
from django.core.management.base import BaseCommand
import logging
import random

from django.db.models import Q, Sum, TextField
from django.db.models.functions import Coalesce, Cast
from openpyxl.workbook import Workbook

from ams import views as attendance_view
from ams.models import LeaveBalance, Attendance, Leave, LeaveBalanceHistory, Break
from ams.views import create_att_calender, calculate_leave_balance, return_back_the_leaves_to_employee, \
    check_if_day_is_endofmonth_for_auto_approval_custom_seed_one
from appraisal.models import Parameter, EmpAppraisal
from erf.models import ERF
from faq.models import FAQ
from hrms import settings
from ijp.models import IJP
from interview.models import FieldsOfExperience, Interviewer, Candidate, Application, InternAnswer
from interview_test.models import TestQA, Test, TestSection
from interview_test.views import get_cur_section_question_mark_pending_if_any
from mapping.models import Designation, Department, EmpData, Profile, Category, HrmsPermissionGroup, Other, \
    MiscellaneousMiniFields
from onboarding.models import Onboard, Salary
from payroll.models import Payroll, VariablePay, UpdateSalaryHistory
from po.models import SupplierAdministration, BillAdministration, BillItemDescription
from team.models import Process, Team
from ticketing.models import Ticket
from utils.handle_permissions import get_handle_permissions
import pandas as pd

from django.db import models

from utils.utils import read_file_format_output, update_my_team, evaluate_essay_by_chat_gpt_in_json_format, \
    get_full_name, get_formatted_name, format_hour_min_second, get_careers_email_message, create_profile_in_qms, \
    update_profile_in_qms, get_decrypted_name, get_dates_range, get_month_number, get_emp_by_emp_id, get_date_value, \
    get_date_time_value, check_if_date_gte_25, non_attrition_profiles, get_attrition_profiles_from_1st_last_month, \
    update_qms_profiles, return_error_response, check_server_status, get_1st_date, get_excluded_dates, update_request, \
    get_decrypted_value, get_column_names, get_gpt, validate_answers, get_group_concat_query_part, \
    create_qms3_profiles, update_profile_in_qms3, create_update_dependencies

# python manage.py seed --mode=refresh

""" Clear all data and creates addresses """
MODE_REFRESH = 'refresh'

""" Clear all data and do not create any object """
MODE_CLEAR = 'clear'
logger = logging.getLogger(__name__)
tz = pytz.timezone(settings.TIME_ZONE)


class Command(BaseCommand):
    help = "seed database for testing and development."

    def add_arguments(self, parser):
        parser.add_argument('--mode', type=str, help="Mode")

    def handle(self, *args, **options):
        self.stdout.write('seeding data...')
        self.stdout.write("here..")
        run_seed(self, options['mode'])
        self.stdout.write('done.')


def create_fields_of_experience():
    FIELDS_OF_EXPERIENCE = {'Human Resources': ["Payroll", "Employee Relations", "HR"],
                            'Talent Acquisition': ["Recruiter", "Onboarding"],
                            'Administration': ["Transport", "Security", "Electrical", "House Keeping", "Plumbing"],
                            'Information Technology': ["Hardware", "Networking", "DataBase", "Dialer",
                                                       "System Administrator"],
                            'Operations': ["Chat", "Email", "Inbound", "Outbound", "Sales", "Business Development",
                                           "CCTV"],
                            'Quality': ["Voice", "Non Voice", "Quality Analysis"],
                            'Learning and Development': ["Trainer"],
                            'Management': ["Team Leader", "Manager", "Assistant Manager", "Senior Manager"],
                            "Analytics": ["Data Analysis", "Data Scientist", "Power BI"],
                            "Development": ["Frontend", "Backend", "Full Stack", "Software Testing"]}

    for key, value in FIELDS_OF_EXPERIENCE.items():
        for i in value:
            FieldsOfExperience.objects.create(category=key, subcategory=i)


# from interview.models import GENDER_CHOICES, QUALF_CHOICES, CANDIDATE_STATUS_CHOICES, SOURCE_CHOICES, LANG_CHOICES

def split_address():
    file_name = "personal_data.xlsx"
    data, workbook, sheet = read_file_format_output(file_name, return_wb_ws=True)
    present_address = data["emp_present_address"]
    pincodes = []
    citys = []
    add1 = []
    for address in present_address:
        city = None
        pincode1 = None
        add1_value = address
        if address and not str(address).isdigit():
            last1 = address.split("-")[-1]
            if len(last1) > 6:
                pincode1 = str(last1.split(" ")[-1]).strip()
            else:
                pincode1 = last1

            pincode1 = pincode1 if str(pincode1).isdigit() else None
            if pincode1:
                city = str(address.split(" ")[-2]).strip()
                city = city if not str(city).isdigit() else None
                if city:
                    add1_value = ", ".join(i for i in address.split(" ")[:-2] if str(i).strip() != "")
                else:
                    add1_value = ", ".join(i for i in address.split(" ")[:-1] if str(i).strip() != "")
            else:
                add1_value = address
        if pincode1:
            citys.append(city)
            pincodes.append(pincode1)
        else:
            citys.append(None)
            pincodes.append(None)
        add1.append(add1_value)
    columns_data = {"temp_city": citys, "temp_add1": add1, "temp_pincode": pincodes}
    last_column_index = sheet.max_column

    for col_index, (header, data) in enumerate(columns_data.items(), start=1):
        sheet.insert_cols(last_column_index + col_index)
        sheet.cell(row=1, column=last_column_index + col_index, value=header)
        for row_index, value in enumerate(data, start=2):
            sheet.cell(row=row_index, column=last_column_index + col_index, value=value)
    workbook.save(file_name)


def update_applied_leaves():
    applied_leave_data = read_file_format_output("Variable Pay Template.xlsx")
    emp_ids = applied_leave_data["emp_id"]
    count = 0
    for emp_id in emp_ids:
        if emp_id is None:
            print("emp_id is None")
            count += 1
            continue
        applied_on = tz.localize(datetime.fromisoformat(applied_leave_data["applied_date"][count].strip()))
        start_date = applied_leave_data["start_date"][count].strip()
        end_date = applied_leave_data["end_date"][count].strip()
        no_of_days = applied_leave_data["no_days"][count].strip()
        agent_reason = applied_leave_data["agent_reason"][count].strip()
        approved_by1 = None
        tl_status = "Approved" if str(applied_leave_data["tl_approval"][count]).strip() == "1" else "Pending"
        profile = Profile.objects.filter(emp_id=emp_id.strip()).last()
        if profile is None:
            count += 1
            print(emp_id, "profile doesnot exist ")
            continue
        tl_comments = None
        if tl_status == "Approved":
            tl_comments = applied_leave_data["tl_reason"][count].strip()
            approved_by1 = profile.team.manager
        leave = Leave.objects.filter(
            start_date=start_date, end_date=end_date,
            no_of_days=no_of_days, reason=agent_reason, profile=profile).last()
        if leave is None:
            print('leave is None')
        leave.applied_on = applied_on
        leave.leave_type = "PL"
        leave.save()
        count += 1


def get_row_data(cols, data, index):
    row_data = {}
    for key in cols:
        if key == "lwd":
            continue

        decrypt_columns = ["fixed_gross", "new_basic_salary", "new_da", "new_oa", "new_ta",
                           "new_fixed_salary", "basic", "da", "new_qa", "ta", "b_plus_d", "total_fixed_gross",
                           "pay_check", "present", "wo", "el", "sl", "cl", "ml", "trng", "hd", "nh", "absent", "att",
                           "na", "total_days", "calendar_days", "paid_days", "lop_days", "el_accrual", "pf_gross",
                           "basic_salary", "da_1", "oa", "ta_1", "statutory_bonus", "variable_pay", "process_incentive",
                           "client_incentive", "bb_rent", "overtime", "night_shift_allowance", "referral", "arrears",
                           "attendance_bonus", "total_gross_salary", "pf_emp", "pf_emr", "esi", "pt", "tds",
                           "cab_deduction", "other_deduction", "notice_period_recovery", "total_ded", "net_salary",
                           "uan", "pf_no", "esic_no"]

        row_data[key] = data[key][index]

        if key in decrypt_columns:
            if row_data[key] is None or str(row_data[key]).strip() == "":
                row_data[key] = 0
            if row_data[key] in ["o", "0", "-", "NA"]:
                row_data[key] = 0
    return row_data


def custom_cr_sh():
    attendance_list = []
    profile = Profile.objects.filter(emp_id=11190).last()
    if profile:
        # yesterday = (datetime.today() - timedelta(hours=24)).date()
        today = datetime.today().date()
        yesterday = (datetime.today() - timedelta(hours=24)).date()
        date1 = datetime.today().replace(year=2023, month=12, day=1).date()
        attendance_list = create_att_calender(profile, attendance_list, date1, today)
        if len(attendance_list) > 0:
            Attendance.objects.bulk_create(attendance_list)
    else:
        print("no profile exists")


def split_name(name):
    name_split = str(name).split(" ")
    first_name = middle_name = last_name = ""
    if len(name_split) == 1:
        first_name = name_split[0]
    elif len(name_split) == 2:
        first_name = name_split[0]
        last_name = name_split[1]
        # if len(first_name) == 1:
        #     first_name = first_name + " " + last_name

    elif len(name_split) >= 3:
        first_name = name_split[0]
        last_name = name_split[-1]
        middle_name = " ".join(i for i in name_split[1:-1])
    return first_name, middle_name, last_name


def create_inactive_profiles(profile_file):
    profiles_data = read_file_format_output(profile_file)
    # print("profiles_data = {0}".format(profiles_data))
    count = 0
    c_team = profiles_data["Process"]
    # print(profiles_data.keys())
    missing_rows = []
    status_exists = False
    all_ok = True
    for base_team_name in c_team:
        # print("prof={0}".format(base_team_name))
        # base_team_name = profiles_data["Core team"][count]
        emp_id = profiles_data["Emp ID"][count]
        if not emp_id:
            continue
        emp_id = str(emp_id).strip()
        emp_name = profiles_data["Emp Name"][count]
        designation_name = profiles_data["Designation"][count]
        team_name = profiles_data["Tasks/Subteam/lob/Group"][count]
        if team_name:
            team_name = str(team_name)
        if base_team_name:
            base_team_name = str(base_team_name)

        emp_status = profiles_data["Status"][count]
        manager = None
        department_name = profiles_data["Department"][count]
        if designation_name:
            designation_name = str(designation_name).strip()
        if department_name:
            department_name = str(department_name).strip()
        department = Department.objects.filter(name__iexact=department_name).last()
        designation = Designation.objects.filter(name__iexact=designation_name).last()

        date_of_joining = profiles_data["DOJ"][count]

        if date_of_joining:
            try:
                date_of_joining = date_of_joining.date()
            except Exception as e:
                date_of_joining = datetime.fromisoformat(date_of_joining)
        if not emp_id or not base_team_name or not team_name or not emp_name or not department or not designation or not emp_status:
            missing_rows.append(
                [profile_file, base_team_name, emp_id, emp_name, designation_name, team_name, emp_status,
                 date_of_joining, department_name])
            print("missing data emp_id", emp_id, "emp_name", emp_name)
            all_ok = False
            count += 1
            continue

        base_team, created = Process.objects.get_or_create(name=base_team_name, defaults={"name": base_team_name,
                                                                                          "department": designation.department,
                                                                                          "is_active": False})
        team, created = Team.objects.get_or_create(base_team=base_team, name=team_name,
                                                   defaults={"name": team_name, "base_team": base_team,
                                                             "manager": manager,
                                                             "is_active": False})
        emp_name = str(emp_name).strip()
        first_name, middle_name, last_name = split_name(emp_name)
        profile, created = Profile.objects.get_or_create(emp_id=emp_id,
                                                         defaults={"emp_id": emp_id, "first_name": first_name,
                                                                   "last_name": last_name,
                                                                   "is_active": False,
                                                                   "date_of_joining": date_of_joining,
                                                                   "status": emp_status,
                                                                   "middle_name": middle_name,
                                                                   "designation": designation,
                                                                   "department": department, "username": emp_id,
                                                                   "team": team, "is_password_changed": True})
        if created:
            profile.set_password("Ecpl@" + str(emp_id))

        if not created:
            profile.date_of_joining = date_of_joining
            profile.first_name = first_name
            profile.last_name = last_name
            profile.middle_name = middle_name
            profile.is_active = False
            profile.designation = designation
            profile.status = emp_status
            profile.team = team
            profile.is_password_changed = True
        profile.save()
        count += 1
    return missing_rows, all_ok


def create_profile(profile_file):
    profiles_data = read_file_format_output(profile_file)
    # print("profiles_data = {0}".format(profiles_data))
    count = 0
    c_team = profiles_data["Process"]
    # print(profiles_data.keys())
    missing_rows = []
    status_exists = False
    all_ok = True
    if "Status" in profiles_data.keys():
        status_exists = True
    for base_team_name in c_team:
        # print("prof={0}".format(base_team_name))
        # base_team_name = profiles_data["Core team"][count]
        emp_id = profiles_data["Emp ID"][count]
        if not emp_id:
            continue
        emp_id = str(emp_id).strip()
        location = profiles_data["Location"][count]
        if location:
            location = str(location).strip()
        emp_name = profiles_data["Emp Name"][count]
        designation_name = profiles_data["Designation"][count]
        team_name = profiles_data["Tasks/Subteam/lob/Group"][count]
        if team_name:
            team_name = str(team_name)
        if base_team_name:
            base_team_name = str(base_team_name)
        if status_exists:
            emp_status = profiles_data["Status"][count]
            if not emp_status:
                emp_status = "Active"
        else:
            emp_status = "Active"
        manager_emp_id = None
        # manager_emp_id = profiles_data["RM Emp ID"][count]
        manager = None
        department_name = profiles_data["Department"][count]
        if designation_name:
            designation_name = str(designation_name).strip()
        if department_name:
            department_name = str(department_name).strip()
        department = Department.objects.filter(name__iexact=department_name).last()
        designation = Designation.objects.filter(name__iexact=designation_name).last()

        date_of_joining = profiles_data["DOJ"][count]
        email = profiles_data["Email"][count]

        if not location or location == "":
            location = ""
        if email:
            email = str(email).strip()
            if email in ["NA", "N/A"]:
                email = ""
        if date_of_joining:
            date_of_joining = date_of_joining.date()
        if not emp_id or not base_team_name or not team_name or not emp_name or not department or not designation:
            missing_rows.append(
                [profile_file, base_team_name, emp_id, emp_name, designation_name, team_name, emp_status, email,
                 date_of_joining, location, department_name])
            print("missing data emp_id", emp_id, "emp_name", emp_name)
            all_ok = False
            count += 1
            continue

        # if manager_emp_id:
        #     manager = Profile.objects.filter(emp_id=manager).last()
        base_team, created = Process.objects.get_or_create(name=base_team_name, defaults={"name": base_team_name,
                                                                                          "department": designation.department,
                                                                                          "is_active": True})
        team, created = Team.objects.get_or_create(base_team=base_team, name=team_name,
                                                   defaults={"name": team_name, "base_team": base_team,
                                                             "manager": manager})
        emp_name = str(emp_name).strip()
        first_name, middle_name, last_name = split_name(emp_name)
        profile, created = Profile.objects.get_or_create(emp_id=emp_id,
                                                         defaults={"emp_id": emp_id, "first_name": first_name,
                                                                   "last_name": last_name,
                                                                   "location": location,
                                                                   "date_of_joining": date_of_joining,
                                                                   "email": email,
                                                                   "status": emp_status,
                                                                   "middle_name": middle_name,
                                                                   "designation": designation,
                                                                   "department": department, "username": emp_id,
                                                                   "team": team, "is_password_changed": True})
        if created:
            profile.set_password("Ecpl@" + str(emp_id))

        print("profile", profile)
        if emp_status != "Active":
            profile.is_active = False
        if not created:
            profile.date_of_joining = date_of_joining
            profile.location = location
            profile.first_name = first_name
            profile.last_name = last_name
            profile.middle_name = middle_name
            profile.designation = designation
            profile.status = emp_status
            profile.email = email
            profile.team = team
            profile.is_password_changed = True
        profile.save()
        count += 1
    return missing_rows, all_ok


def add_filename_to_list(filename, file_name_list):
    with open(file_name_list, 'a') as file:
        file.write(filename + '\n')


def filename_exists_in_list(filename, file_name_list):
    with open(file_name_list, 'r') as file:
        file_names = file.read().splitlines()
    return filename in file_names


def get_approved_by(app_by, nm1, nm2, nm3, rm1, rm2, rm3):
    if app_by == nm1:
        return rm1
    elif app_by == nm2:
        return rm2
    elif app_by == nm3:
        return rm3
    else:
        return rm1


def create_leave_balance(balance_file):
    balance_data = read_file_format_output(balance_file)
    count = 0
    profile_not_exist = []
    emp_id = balance_data["emp_id"]
    for emp in emp_id:
        if not emp:
            count += 1
            continue
        employee = Profile.objects.filter(emp_id=emp).last()
        if not employee:
            print("profile doesn't exist = {0}".format(emp))
            profile_not_exist.append(emp)
            count += 1
            continue
        paid_leaves = balance_data["pl_balance"][count]
        paid_leaves = round(float(paid_leaves), 2) if paid_leaves else 0

        sick_leaves = balance_data["sl_balance"][count]
        sick_leaves = round(float(sick_leaves), 2) if sick_leaves else 0
        print("emp_id = {0}, employee = {1}, pl = {2}, sl = {3}, count = {4}".format(emp, employee, paid_leaves,
                                                                                     sick_leaves, count))
        total = float(paid_leaves + sick_leaves)
        leave, created = LeaveBalance.objects.get_or_create(profile=employee,
                                                            defaults={"profile": employee, "sick_leaves": sick_leaves,
                                                                      "paid_leaves": paid_leaves, "total": total,
                                                                      "updated_at": datetime.now()})
        if not created:
            leave.sick_leaves = sick_leaves
            leave.paid_leaves = paid_leaves
            leave.total = total
            leave.updated_at = datetime.now()
        leave.save()
        count += 1
    return profile_not_exist


def get_att_obj(date1, profile):
    return Attendance(status="Scheduled", date=date1, end_time="09:00:00",
                      total_time="00:00:00", is_training=False, profile=profile,
                      is_self_marked=False, is_att_cor_req=False)


def import_calender(file_name):
    with open(file_name, "r") as file:
        csv_reader = csv.DictReader(file)
        count = 0
        profiles_dict = {}
        att_list = []
        for cal_data in csv_reader:
            date1 = cal_data["date"]
            att_actual = cal_data["att_actual"]

            if not att_actual or att_actual == "" or date1 is None or str(date1) == "":
                count += 1
                continue
            approved_by = cal_data["appoved_by"]
            approved_at = cal_data["approved_on"]
            emp_id = cal_data["emp_id"]
            if not emp_id or str(emp_id) == "":
                continue
            if emp_id not in profiles_dict.keys():
                profile = Profile.objects.filter(emp_id=emp_id).last()
                if profile is None:
                    continue
                profiles_dict[emp_id] = profile
            else:
                profile = profiles_dict[emp_id]
            if profile is None:
                continue
            if approved_at:
                approved_at = tz.localize(datetime.fromisoformat(approved_at))
            else:
                approved_at = None
            nm1 = cal_data["rm1"]
            nm2 = cal_data["rm1"]
            nm3 = cal_data["rm1"]
            rm1 = cal_data["rm1_id"]
            rm2 = cal_data["rm2_id"]
            rm3 = cal_data["rm3_id"]

            app_by_emp_id = get_approved_by(approved_by, nm1, nm2, nm3, rm1, rm2, rm3)
            if app_by_emp_id in profiles_dict.keys():
                approved_by = profiles_dict[app_by_emp_id]
            else:
                # print("it came here first")
                approved_by = Profile.objects.filter(emp_id=app_by_emp_id).last()
                if approved_by is None:
                    # print("it came here second")

                    approved_by = Profile.objects.filter(emp_id=rm1).last()
                    if approved_by is None:
                        # print("it came here third")

                        approved_by = Profile.objects.filter(emp_id=rm2).last()
                        if approved_by is None:
                            # print("it came here fourth")

                            approved_by = Profile.objects.filter(emp_id=rm3).last()

                            if approved_by is None:
                                # print("it came here fifth")

                                # print(cal_data)
                                if "6043" in profiles_dict.keys():
                                    approved_by = profiles_dict["6043"]
                profiles_dict[app_by_emp_id] = approved_by
            total_time = None
            is_training = False
            att_actual = str(att_actual).lower()
            is_self_marked = False

            if att_actual == "present":
                is_self_marked = True
                total_time = "09:00:00"

            if att_actual == "training":
                is_training = True
            if att_actual == "unmarked":
                att_actual = "Scheduled"
            if len(att_actual) <= 4:
                att_actual = str(att_actual).upper()
            else:
                att_actual = " ".join(i.capitalize() for i in att_actual.split())

            count += 1
            # print(approved_by)
            att_list.append(
                Attendance(status=att_actual, date=date1, end_time="09:00:00",
                           total_time=total_time, is_training=is_training, sch_upd_by=approved_by, profile=profile,
                           is_self_marked=is_self_marked, approved_at=approved_at, approved_by=approved_by,
                           is_att_cor_req=False))
        if len(att_list) > 0:
            Attendance.objects.bulk_create(att_list)


def create_employees(folder_name, missing_file_name, file_name_list):
    with os.scandir(folder_name) as it:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        file_name_list = file_name_list
        if not os.path.exists(file_name_list):
            # Create the file_name_list file if it doesn't exist
            open(file_name_list, 'w').close()

        for entry in it:
            if entry.name.endswith(".xlsx") and entry.is_file():
                # print(entry.name, entry.path)
                filename = entry.name
                if filename_exists_in_list(filename, file_name_list) and filename != "missing_data.xlsx":
                    print(f"{filename} already exists in {file_name_list}.")
                else:
                    missing_rows, all_ok = create_profile(entry.path)
                    # Add the filename to the file_name_list file
                    if len(missing_rows) > 0:
                        for row in missing_rows:
                            worksheet.append(row)
                        workbook.save(missing_file_name)
                    if filename != missing_file_name and all_ok:
                        add_filename_to_list(filename, file_name_list)
                    print(f"Added {filename} to {file_name_list}.")

        workbook.close()


def update_erf_created_by(file_name):
    erf_data = read_file_format_output(file_name)
    process_data = erf_data["campaign"]
    count = 0
    erf_list = []
    for process_name in process_data:
        created_by_id = erf_data["created_by_id"][count]
        erf_id = erf_data["id"][count]
        if erf_id:

            erf = ERF.objects.filter(id=erf_id).last()
            if created_by_id:
                created_by_id = str(created_by_id).strip()
                created_by = Profile.objects.filter(emp_id=created_by_id).last()
                erf.created_by = created_by
                erf_list.append(erf)
        count += 1
    if len(erf_list) > 0:
        ERF.objects.bulk_update(erf_list, ["created_by"])


def decrypt_payslip_encrypted(file_name):
    pay_slip_data = read_file_format_output(file_name)
    decrypt_columns = ["lwd", "bank_name", "account_number", "ifsc_code", "len", "new_pay_grade",
                       "new_fixed_salary", "basic", "da", "new_qa", "ta", "b_plus_d", "total_fixed_gross",
                       "pay_check", "present", "wo", "el", "sl", "cl", "ml", "trng", "hd", "nh", "absent", "att",
                       "na", "total_days", "calendar_days", "paid_days", "lop_days", "el_accrual", "pf_gross",
                       "basic_salary", "da_1", "oa", "ta_1", "statutory_bonus", "variable_pay", "process_incentive",
                       "client_incentive", "bb_rent", "overtime", "night_shift_allowance", "referral", "arrears",
                       "attendance_bonus", "total_gross_salary", "pf_emp", "pf_emr", "esi", "pt", "tds",
                       "cab_deduction", "other_deduction", "notice_period_recovery", "total_ded", "net_salary",
                       "uan", "pf_no", "esic_no"]
    decrypted_data = {}
    for key in pay_slip_data:
        decrypted_data[key] = []
        if key in decrypt_columns:
            for val in pay_slip_data[key]:
                decrypted_data[key].append(get_decrypted_name(val))
        else:
            decrypted_data[key] = pay_slip_data[key]
    wb = Workbook()
    ws = wb.active
    headers = list(pay_slip_data.keys())
    ws.append(list(headers))
    for row in zip(*decrypted_data.values()):
        ws.append(row)
    excel_file_name = 'pay_slip_decrypted.xlsx'
    wb.save(excel_file_name)


def import_erf(file_name):
    erf_data = read_file_format_output(file_name)
    process_data = erf_data["campaign"]
    count = 0
    erf_list = []
    for process_name in process_data:
        if not process_name:
            count += 1
            continue
        ticket_id = erf_data["ticket_id"][count]

        designation_name = str(erf_data["designation"][count]).strip()
        designation = Designation.objects.filter(name__iexact=designation_name).last()
        if designation_name is None or str(designation_name) == "":
            print("invalid designation, name", designation_name, "existing desi", ticket_id)
            count += 1
            continue
        department_name = erf_data["department"][count]
        if designation is None:
            print("invalid designation", designation_name, department_name)
            break

        process = Process.objects.filter(name__iexact=process_name).last()

        if process is None:
            process = Process.objects.create(name=process_name, department=designation.department)
        process_type_one = erf_data["process_type_one"][count]
        process_type_two = erf_data["process_type_two"][count]
        process_type_three = erf_data["process_type_three"][count]
        salary_range_frm = erf_data["salary_rang_frm"][count]
        salary_range_to = erf_data["salary_rang_to"][count]
        if not str(salary_range_frm).isdecimal() or not str(salary_range_to).isdecimal():
            count += 1
            print("invalid salary range frm/to  {0}".format(ticket_id))
            continue

        requisition_date = erf_data["requisition_date"][count]
        updated_at = datetime.fromisoformat(erf_data["edited_date"][count])
        updated_at = tz.localize(updated_at)
        requisition_date = datetime.fromisoformat(requisition_date)
        created_at = requisition_date
        created_at = tz.localize(created_at)
        hc_req = erf_data["Head Count_required"][count]
        created_by_mgr = erf_data["created_by_manager_id"][count]
        if created_by_mgr:
            created_by_mgr = str(created_by_mgr).strip()
        created_by = Profile.objects.filter(emp_id=created_by_mgr).last()
        if created_by is None:
            print("manager doesn't exist {0} {1}".format(process_name, created_by_mgr))
            # count += 1
            # continue
        qualification = erf_data["qualification"][count]
        if qualification == "12th":
            qualification = "Pre University"
        elif qualification == "SSLC":
            qualification = "Matriculation"
        other_qualification = erf_data["skills_set"][count]
        languages = erf_data["languages"][count]
        if languages:
            languages = languages[1:-1].replace("'", "")

        shift_timing = erf_data["shift_timing"][count]
        shift_timing_frm = erf_data["shift_timing_frm"][count]
        shift_timing_to = erf_data["shift_timing_to"][count]
        shift_timing = str(shift_timing.split("-")[0]).strip() if shift_timing else None
        if str(shift_timing).lower() in ["day", "night"]:
            shift_timing = "Fixed"
        if not shift_timing:
            print("no shift timing {0} {1}".format(process_name, ticket_id))
            count += 1
            continue
        type_of_working = erf_data["type_of_working"][count]
        if type_of_working is None:
            print("no type of working {0} {1}".format(process_name, ticket_id))
            count += 1
            continue
        week_frm = erf_data["week_from"][count]
        week_to = erf_data["week_to"][count]
        week_offs = f"{week_frm}{week_to}"
        requisition_type = erf_data["requisition_type"][count]
        reason_for_replace = erf_data["reason_for_replace"][count]
        closure_date = erf_data["closure_date"][count]
        dead_line = erf_data["dead_line"][count]
        is_open = True
        closed_by_id = erf_data["closed_by_id"][count]
        closed_by = None
        if closed_by_id:
            closed_by = Profile.objects.filter(emp_id=closed_by_id).last()
            is_open = False
        mrv_status = erf_data["Mervyn _Approval"][count]
        if mrv_status == "Approved":
            approval_status = "Approved"
            approved_by = Profile.objects.filter(emp_id=6043).last()

        fields_of_experience = list(
            FieldsOfExperience.objects.filter(category__iexact=designation.department.name).values_list(
                "subcategory", flat=True))
        fields_of_experience = ",".join(i for i in fields_of_experience)

        no_of_rounds = 1
        round_names = "Round 1"
        interview_persons = created_by_mgr
        erf_id = erf_data["id"][count]
        erf = ERF.objects.filter(id=erf_id).last()
        if erf:
            erf.updated_at = updated_at
            erf.created_at = created_at
            erf.week_offs = week_offs
            erf.shift_timing = shift_timing
            erf.save()
        if not erf:
            erf_list.append(ERF(id=erf_id, approval_status=approval_status, approved_or_rejected_by=approved_by,
                                requisition_date=requisition_date, hc_req=hc_req, created_by=created_by,
                                process=process, department=designation.department, designation=designation,
                                process_type_one=process_type_one, process_type_two=process_type_two,
                                process_type_three=process_type_three, pricing=0, salary_range_frm=salary_range_frm,
                                salary_range_to=salary_range_to, qualification=qualification,
                                other_quali=other_qualification, fields_of_experience=fields_of_experience,
                                languages=languages, shift_timing=shift_timing, shift_timing_frm=shift_timing_frm,
                                shift_timing_to=shift_timing_to, type_of_working=type_of_working,
                                week_offs=week_offs,
                                requisition_type=requisition_type, reason_for_replace=reason_for_replace,
                                no_of_rounds=no_of_rounds, round_names=round_names,
                                interview_persons=interview_persons,
                                closure_date=closure_date, deadline=dead_line, closed_by=closed_by,
                                is_erf_open=is_open, updated_at=updated_at,
                                always_open=False, created_at=created_at))
        count += 1
    if len(erf_list) > 0:
        ERF.objects.bulk_create(erf_list)


def update_team(file_name):
    missing_team_file = "missing_team.xlsx"
    profiles_data = read_file_format_output(file_name)
    if not os.path.exists(missing_team_file):
        workbook = openpyxl.Workbook()
    else:
        workbook = openpyxl.load_workbook(missing_team_file)
    worksheet = workbook.active

    count = 0
    team_data = profiles_data["Team"]
    rows = []
    for team_name in team_data:
        if team_name:
            team_name = str(team_name).strip()
        team = Team.objects.filter(name=team_name).last()
        emp_id = profiles_data["EMP ID"][count]
        if emp_id:
            emp_id = str(emp_id).strip()
        manager = Profile.objects.filter(emp_id=emp_id).last()
        if team is None or manager is None:
            # print("team with team_name {0} not updated team = {1} manager = {2}".format(team_name, team, manager))
            print("not exists {0}, emp_id {1}".format(team_name, emp_id))
            rows.append([team_name, emp_id, team.name if team else "", manager.emp_id if manager else ""])
            count += 1
            continue
        if team:
            team.manager = manager
            team.save()
        count += 1
    if len(rows) > 0:
        for row in rows:
            worksheet.append(row)
        workbook.save(missing_team_file)
    workbook.close()


def update_dob_in_profile():
    file_name = "DOBS.xlsx"
    dob_data = read_file_format_output(file_name)
    count = 0
    emp_id = dob_data["emp_id"]
    for emp in emp_id:
        if not emp:
            count += 1
            continue
        employee = Profile.objects.filter(emp_id=str(emp).strip()).last()
        if not employee:
            print("profile doesn't exist = {0}".format(emp))
            count += 1
            continue
        dob = dob_data["DOB"][count]
        if not dob:
            count += 1
            continue
        try:

            dob = dob.date()
        except Exception as e:
            try:
                dob = datetime.fromisoformat(dob)
            except Exception as e:
                print(str(e), dob)
                count += 1
                break
                continue
        val = datetime.today().date() - dob
        year = int(val.days / 365)
        if year >= 100 or year < 18:
            count += 1
            continue
        employee.dob = dob
        employee.save()
        count += 1


def migrate_payslip():
    map_columns = {"new_fixed_salary": "fixed_gross",
                   "basic_salary": "new_basic_salary",
                   "da_1": "new_da",
                   "oa": "new_oa",
                   "ta_1": "new_ta",
                   "location": "branch",
                   "emp_process": "process",
                   "name_as_per_account_number": "name_as_per_bank_acc",
                   "new_pay_grade": "pay_grade",
                   "emp_desi": "designation",
                   "emp_department": "department"}

    headers = ["year", "month", "month_number", "emp_id", "emp_name", "doj", "location", "lwd", "status",
               "gender", "emp_process", "emp_desi", "emp_department", "payment_mode", "bank_name",
               "name_as_per_account_number", "account_number", "ifsc_code", "len", "new_pay_grade",
               "new_fixed_salary", "basic", "da", "new_qa", "ta", "b_plus_d", "total_fixed_gross", "pay_check",
               "present", "wo", "el", "sl", "cl", "ml", "trng", "hd", "nh", "absent", "att", "na", "total_days",
               "calendar_days", "paid_days", "lop_days", "el_accrual", "pf_gross", "basic_salary", "da_1", "oa",
               "ta_1", "statutory_bonus", "variable_pay", "process_incentive", "client_incentive", "bb_rent",
               "overtime", "night_shift_allowance", "referral", "arrears", "attendance_bonus", "total_gross_salary",
               "pf_emp", "pf_emr", "esi", "pt", "tds", "cab_deduction", "other_deduction", "notice_period_recovery",
               "total_ded", "net_salary", "uan", "pf_no", "esic_no"]
    payslip_data = read_file_format_output("PaySlip-2023-11-23_decrypted_formatted.xlsx", skip_cols=["len", "lwd"],
                                           map_columns=map_columns)
    emp_ids = payslip_data["emp_id"]
    payroll_data = []
    profiles_dict = {}
    count = 0
    headers = payslip_data.keys()
    for emp_id in emp_ids:
        if emp_id not in profiles_dict:
            profiles_dict[emp_id] = Profile.objects.filter(emp_id=emp_id).last()
        profile = profiles_dict[emp_id]
        if profile is None:
            # print("emp_id {0} not found".format(emp_id))
            count += 1
            continue
        payroll_data.append(Payroll(profile=profile, **get_row_data(headers, payslip_data, count)))
        count += 1
    Payroll.objects.bulk_create(payroll_data)


def correct_lwd():
    profiles = Profile.objects.filter(is_active=False, is_superuser=0, last_working_day__isnull=False)
    for profile in profiles:
        if profile.status == "Attrition":
            if profile.last_working_day >= datetime.today().replace(month=11, year=2023, day=1).date():
                attendance = Attendance.objects.filter(profile=profile, date=profile.last_working_day).last()
                if attendance:
                    if attendance.status == "Attrition":
                        print(profile)
                        profile.last_working_day = profile.last_working_day - timedelta(days=1)
                        profile.save()
        elif profile.status in ["NCNS", "Bench"]:
            profile.last_working_day = None
            profile.save()


def create_empty_leave_table():
    profiles = Profile.objects.filter(is_superuser=0)
    cr_leave_balance_records = []
    for profile in profiles:
        leave_balance = LeaveBalance.objects.filter(profile=profile).last()
        if leave_balance is None:
            cr_leave_balance_records.append(LeaveBalance(profile=profile, sick_leaves=0, paid_leaves=0, total=0))
    if len(cr_leave_balance_records) > 0:
        LeaveBalance.objects.bulk_create(cr_leave_balance_records)


def correct_sh():
    date1 = datetime.now().replace(year=2023, month=11, day=1).date()
    attendances = Attendance.objects.filter(date__month=date1.month, date__year=date1.year)
    profiles = Profile.objects.all()
    date_range = get_dates_range(date1, date1.replace(day=30))
    for profile in profiles:
        for date2 in date_range:
            pro_attendances = attendances.filter(profile=profile, date=date2)
            if pro_attendances.count() > 1:
                working_count = 0
                for att in pro_attendances:
                    print(att.status, att.date, att.profile.emp_id)
                    # if working_count == 0 and att.status == "Scheduled":
                    #     working_count = 1
                    # att.delete()


def correct_erf():
    erfs = ERF.objects.filter(no_of_rounds=0)
    for erf in erfs:
        no_of_rounds = len(erf.interview_persons.split(","))
        if no_of_rounds > 0:
            erf.no_of_rounds = no_of_rounds
            erf.save()


def set_full_name_dob():
    profiles = Profile.objects.all()
    for profile in profiles:
        profile.full_name = get_formatted_name(profile)
        profile.dob = profile.onboard.dob if profile.onboard else None
        profile.save()
    candidates = Candidate.objects.all()
    for candidate in candidates:
        candidate.full_name = get_formatted_name(candidate)
        candidate.save()


def update_total_time():
    date1 = datetime.today().replace(year=2023, month=11, day=1).date()
    attendances = Attendance.objects.filter(date__gte=date1)
    att_list = []
    for att in attendances:
        if att.total_time:
            tt = att.total_time.split(":")
            h, m, s = format_hour_min_second(tt[0], tt[1], tt[2])
            att.total_time = f"{h}:{m}:{s}"
            att_list.append(att)
    if len(att_list) > 0:
        Attendance.objects.bulk_update(att_list, ["total_time"])


def update_leave_balance_history(file_name):
    lb_data = read_file_format_output(file_name)
    tids = lb_data["id"]
    count = 0
    lbh_list = []
    for tid in tids:
        lbh = LeaveBalanceHistory.objects.filter(id=tid).last()
        lbh.date = tz.localize(get_date_time_value(lb_data["date"][count]))
        lbh_list.append(lbh)
        count += 1
    if len(lbh_list) > 0:
        LeaveBalanceHistory.objects.bulk_update(lbh_list, ['date'])


def import_leave_balance_history(file_name):
    lb_data = read_file_format_output(file_name)
    emp_ids = lb_data["emp_id"]
    emp_id_dict = {}
    count = 0
    lbh_list = []
    for emp_id in emp_ids:
        if emp_id not in emp_id_dict.keys():
            emp_id_dict[emp_id] = Profile.objects.filter(emp_id=emp_id).last()
        emp = emp_id_dict[emp_id]
        if emp is None:
            count += 1
            continue
        date1 = tz.localize(get_date_time_value(lb_data["date"][count]))
        lbh_list.append(LeaveBalanceHistory(profile=emp, leave_type=lb_data["leave_type"][count],
                                            transaction=lb_data["transaction"][count],
                                            no_of_days=lb_data['no_days'][count], date=date1,
                                            total=lb_data["total"][count]))
        count += 1
    if len(lbh_list) > 0:
        LeaveBalanceHistory.objects.bulk_create(lbh_list)


def import_variable_pay(file_name):
    map_columns = {"bonus": "attendance_bonus", "ot": "overtime", "for_year": "year"}
    variable_pay_data = read_file_format_output(file_name, map_columns=map_columns)
    emp_ids = variable_pay_data["emp_id"]
    count = 0
    keys = variable_pay_data.keys()
    vp_list = []
    vp_update = []
    for emp_id in emp_ids:
        for key in keys:
            variable_pay_data[key][count] = str(variable_pay_data[key][count]).strip()
        emp_id = str(emp_id).strip()
        month = get_month_number(variable_pay_data['for_month'][count])

        created_at = get_date_time_value(variable_pay_data['created_on'][count])
        try:
            created_at = tz.localize(created_at)
        except:
            pass
        created_by = variable_pay_data["created_by_id"][count]
        try:
            profile = get_emp_by_emp_id(emp_id)
        except:
            count += 1
            continue
        try:
            created_by = get_emp_by_emp_id(created_by)
        except:
            created_by = None
        year = str(variable_pay_data['year'][count]).strip()
        vp = VariablePay.objects.filter(emp_id=emp_id, year=year, month=month).last()
        vp_data = {'total': variable_pay_data["total"][count],
                   'created_at': created_at, 'month': month,
                   'emp_id': emp_id,
                   'year': variable_pay_data['year'][count], 'comment': variable_pay_data['comments'][count],
                   'process_incentive': variable_pay_data['process_incentive'][count],
                   'client_incentive': variable_pay_data['client_incentive'][count],
                   'overtime': variable_pay_data['overtime'][count],
                   'attendance_bonus': variable_pay_data['attendance_bonus'][count],
                   'created_by': created_by, 'profile': profile}

        if vp is None:
            vp_list.append(
                VariablePay(**vp_data))
        else:
            for key in vp_data.keys():
                setattr(vp, key, vp_data[key])
            vp_update.append(vp)
        count += 1
    if len(vp_list) > 0:
        VariablePay.objects.bulk_create(vp_list)
    if len(vp_update) > 0:
        VariablePay.objects.bulk_update(vp_update,
                                        ['total', 'created_at', 'month', 'year', 'comment', 'process_incentive',
                                         'client_incentive', 'overtime', 'attendance_bonus', 'created_by', 'profile'])


def add_new_year_sl(profiles):
    first_day = datetime.today().date().replace(day=1, year=2025, month=1)
    date_now = tz.localize(datetime.now())
    objs = []
    lb_list = []
    prev_month_last_day = date_now.date().replace(day=1) - timedelta(days=1)
    for profile in profiles:
        sl_leaves = Leave.objects.filter(profile=profile, start_date__gte=first_day, leave_type="SL",
                                         status="Pending")
        approved_sl = Attendance.objects.filter(profile=profile, status="SL", date__gte=first_day).count()
        sl_leaves_count = approved_sl
        for sll in sl_leaves:
            sl_leaves_count += sll.no_of_days
        leave_balance = LeaveBalance.objects.filter(profile=profile).last()
        if sl_leaves_count == 0:
            pl, sl = calculate_leave_balance(profile, prev_month_last_day)

            leave_balance.sick_leaves = sl
            leave_balance.total = round(leave_balance.sick_leaves + leave_balance.paid_leaves, 2)
            lb_list.append(leave_balance)
            objs.append(
                LeaveBalanceHistory(profile=profile, date=date_now, leave_type="SL",
                                    pl_balance=leave_balance.paid_leaves,
                                    sl_balance=leave_balance.sick_leaves,
                                    transaction="{0} SL credited back for new year".format(sl),
                                    no_of_days=sl, total=leave_balance.total))
        else:
            objs.append(LeaveBalanceHistory(profile=profile, date=date_now, leave_type="SL",
                                            pl_balance=leave_balance.paid_leaves,
                                            sl_balance=leave_balance.sick_leaves,
                                            transaction="{0} SL already applied, 0 SL credited back for new year".format(
                                                sl_leaves_count),
                                            no_of_days=0, total=leave_balance.total))

    if len(objs) > 0:
        LeaveBalanceHistory.objects.bulk_create(objs)
    if len(lb_list) > 0:
        LeaveBalance.objects.bulk_update(lb_list, ["sick_leaves", "total"])


def check_and_deduct_lb_and_applied_leaves():
    LeaveBalHist_list = []
    leave_bal_list = []
    date_now = datetime.today().date()
    profiles = Profile.objects.exclude(is_superuser=True)
    for profile in profiles:
        first_day = datetime.today().date().replace(day=1, month=1)
        approved_leave_count = Attendance.objects.filter(profile=profile, date__gte=first_day, status="PL").count()
        applied_leave = Leave.objects.filter(status="Pending", profile=profile, start_date__gte=first_day,
                                             leave_type="PL")
        appl_att_count = 0
        lb, created = LeaveBalance.objects.get_or_create(profile=profile)
        lb.sick_leaves = 0

        for leave in applied_leave:
            appl_att_count += leave.no_of_days
        total_applied_leaves = approved_leave_count + appl_att_count
        if total_applied_leaves >= 10:
            LeaveBalHist_list.append(
                LeaveBalanceHistory(profile=profile, date=date_now, leave_type="SL and PL", pl_balance=0,
                                    sl_balance=0, transaction="Leaves deducted as per new year policy",
                                    no_of_days=lb.total, total=0))
            lb.total = 0
            lb.paid_leaves = 0
        else:
            if lb.paid_leaves + total_applied_leaves > 10:
                deduct = round(lb.paid_leaves + total_applied_leaves - 10, 2)
                lb.paid_leaves = round(lb.paid_leaves - deduct, 2)
                lb.total = lb.paid_leaves
                LeaveBalHist_list.append(LeaveBalanceHistory(profile=profile, date=date_now, leave_type="SL and PL",
                                                             pl_balance=lb.paid_leaves, sl_balance=0,
                                                             no_of_days=deduct, total=lb.total,
                                                             transaction="Leaves deducted as per new year policy",
                                                             ))
            else:
                lb.total = lb.paid_leaves
        leave_bal_list.append(lb)

    if len(LeaveBalHist_list) > 0:
        LeaveBalanceHistory.objects.bulk_create(LeaveBalHist_list)
    if len(leave_bal_list) > 0:
        LeaveBalance.objects.bulk_update(leave_bal_list, ["total", "paid_leaves", "sick_leaves"])


def import_supplier_administration():
    file_name = "SupplierAdministration-2024-01-05.xlsx"
    b_data = read_file_format_output(file_name)
    name_list = b_data["name"]
    i = 0
    for name in name_list:
        SupplierAdministration.objects.create(name=name, contact_person=b_data["cantact_person"][i],
                                              contact_no=b_data["contact_no"][i], address=b_data["address"][i],
                                              contact_email=b_data["contact_email"][i], pan_no=b_data["pan"][i],
                                              gst=b_data["gst"][i], acc_name=b_data["acc_name"][i],
                                              account_number=b_data["acc_no"][i], bank_name=b_data["bank_name"][i],
                                              bank_branch=b_data["bank_branch"][i], ifsc=b_data["ifsc"][i],
                                              cin_code=b_data["cin_code"][i] if b_data["cin_code"][i] else None)
        i += 1


def import_po():
    file_name = "BillAdministration-2024-01-05.xlsx"
    b_data = read_file_format_output(file_name)

    i_data = read_file_format_output("ItemDescriptionAdministration-2024-01-05.xlsx")

    po_no_list = b_data["po_no"]
    item_po_id_list = i_data["bill"]
    i = 0

    for po_no in po_no_list:
        sup_contact_no = b_data["supplier_contact_no"][i]
        sup_contact_email = b_data["supplier_contact_email"][i]

        supplier = SupplierAdministration.objects.filter(
            Q(contact_no=sup_contact_no) | Q(contact_email=sup_contact_email)).last()
        ba = BillAdministration.objects.create(supplier=supplier, project=b_data["project"][i], po_no=po_no,
                                               date=b_data["date"][i], billing_office=b_data["billing_office"][i],
                                               category=b_data["category"][i],
                                               delivery_office=b_data["delivery_office"][i],
                                               contact_person=b_data["contact_person"][i],
                                               contact_no=b_data["contact_no"][i],
                                               contact_email=b_data["contact_email"][i],
                                               terms_and_conditions=b_data["terms_conditions"][i],
                                               amount_in_words=b_data["amount_words"][i], pan=b_data["pan"][i],
                                               gst=b_data["gst"][i], total_amount=b_data["total_amount"][i],
                                               gst_amount=b_data["gst_amount"][i],
                                               grand_total=b_data["grand_total"][i])
        if b_data["id"][i] in item_po_id_list:
            for index, value in enumerate(item_po_id_list):
                if value == b_data["id"][i]:
                    BillItemDescription.objects.create(bill=ba, description=i_data["description"][index],
                                                       quantity=i_data["qty"][index],
                                                       gst_percent=i_data["gst_percent"][index],
                                                       gst_amount=i_data["gst_amount"][index],
                                                       price=i_data["price"][index], amount=i_data["amount"][index])

        i += 1


def import_pay_slip(file_name):
    from payroll.serializers import UploadPayrollSerializer
    xl_data = read_file_format_output(file_name)
    model_fields = Payroll._meta.get_fields()
    column_names = [field.name for field in model_fields if field.concrete]
    column_names.remove("id")
    i = 0
    obj_list = []
    emp_id_list = xl_data.get('emp_id')
    if not emp_id_list or emp_id_list == ():
        raise ValueError("Please add valid EMP-IDs to the file uploaded.")
    payroll_cr_list = []
    while emp_id_list and i < len(emp_id_list):
        emp_id = emp_id_list[i]
        if not emp_id:
            i += 1
            continue
        try:
            employee = Profile.objects.get(emp_id=emp_id)
        except Exception as e:
            raise ValueError("employee with emp_id {0} doesn't exist".format(emp_id))
        if emp_id is None:
            break
        obj_data = {}
        for col_name in column_names:
            # map_name = map_columns.get(col_name, col_name)
            xl_col = xl_data.get(col_name)
            if xl_col is None:
                continue
            value = xl_col[i]
            obj_data[col_name] = value
        doj = obj_data.get('doj')
        if isinstance(doj, datetime):
            obj_data['doj'] = doj.date()
        serializer = UploadPayrollSerializer(data=obj_data)
        if not serializer.is_valid():
            return_error_response(serializer, raise_exception=True)

        pay_obj = Payroll.objects.filter(emp_id=emp_id, month_number=obj_data['month_number'],
                                         year=obj_data["year"]).last()
        if pay_obj:
            for key, value in serializer.validated_data.items():
                setattr(pay_obj, key, value)
                pay_obj.save()
            logger.info("updated {0}".format(emp_id))
        else:
            serializer.validated_data["profile"] = employee
            payroll_cr_list.append(Payroll(**serializer.validated_data))

        i += 1
    if len(payroll_cr_list) > 0:
        Payroll.objects.bulk_create(payroll_cr_list)


def change_password(username, new_password):
    try:
        # Get the user object
        user = Profile.objects.get(username=username)

        # Change the password
        user.set_password(new_password)
        user.save()

        print(f"Password changed successfully for user: {username}")
    except Exception as e:
        print(f"Failed to change password for user '{username}': {e}")


def add_score():
    answers = InternAnswer.objects.filter(score__isnull=True)
    size = 25
    batches = [answers[i:i + size] for i in range(0, len(answers), size)]
    for btch_no, batch in enumerate(batches):
        for ansr in batch:
            question = ansr.question.question_text
            answer = ansr.answer_text
            score = validate_answers(answer, question)
            ansr.score = score
            ansr.save()
        print(f"{(btch_no + 1) * size} is completed out of {len(answers)}")


def run_seed(self, mode, value=None):
    """ Seed database based on mode

    :param mode: refresh / clear
    :return:
    """
    if mode == "force_logout_all":
        from rest_framework.authtoken.models import Token
        Token.objects.all().delete()

    # if mode == "force_logout_inactive_users":
    #     now = datetime.now().replace(microsecond=0)-timedelta(minutes=60)
    #     Profile.objects.filter(last_login__gte)

    if mode == "generate_qrcode":
        otp = pyotp.TOTP("BCCNFT2DIZG3T2XVYVGY5ZM36ETFEELB")
        provisioning_uri = otp.provisioning_uri(name='33434', issuer_name="ECPL HRMS")

        # Generate QR code
        qr = qrcode.make(provisioning_uri)
        qr_code_io = io.BytesIO()
        qr.save(qr_code_io, format='PNG')
        qr_code_io.seek(0)
        print('data:image/png;base64,' + base64.b64encode(qr_code_io.getvalue()).decode(
            'utf-8'))

    if mode == "verify_otp":
        totp = pyotp.TOTP("BCCNFT2DIZG3T2XVYVGY5ZM36ETFEELB")
        otp = 908072
        print(totp.verify(str(otp)))

    if mode == "0011":
        get_handle_permissions().clean_up_permissions()
        get_handle_permissions().add_permissions()

    if mode == "cr_sh":
        attendance_view.generate_attendance_calender()

    if mode == "correct_sh":
        correct_sh()
    if mode == "cor_erf":
        correct_erf()
    # if mode == "1111":
    #     attendance_view.handle_csv_file("Profile-2023-02-22.csv")

    if mode == "mark_absent":
        attendance_view.mark_absent_if_not_worked()
    if mode == "import_calender":
        import_calender("calender_data.csv")

    if mode == "import_erf":
        import_erf("erfs.xlsx")
    if mode == "update_erf":
        update_erf_created_by("erfs.xlsx")
    if mode == "update_team":
        update_team("All_Teams.xlsx")

    if mode == "update_my_team":
        update_my_team()
        update_qms_profiles()
    if mode == "create_inactive_profile":
        missing_rows, all_ok = create_inactive_profiles("inactive_profiles.xlsx")
        print(missing_rows)

    if mode == "create_profile":
        create_employees("profile_data_files", "missing_data.xlsx",
                         "profile_file_names.txt")
    if mode == "upload_leave_balance":
        file_name = "leave_balance.xlsx"
        prof_not_exist = create_leave_balance(file_name)
    if mode == "training_to_present":
        nov_1 = datetime.today().replace(year=2023, month=11, day=1)
        Attendance.objects.filter(status="Training", date__gte=nov_1).update(status="Present", is_training=True)
    if mode == "create_empty_leave_table":
        create_empty_leave_table()
    if mode == "clean_up_erfs":
        erfs = ERF.objects.filter(no_of_rounds=0).update(interview_persons="10323", no_of_rounds=1,
                                                         round_names="Round 1")
    if mode == "set_full_name_dob":
        set_full_name_dob()
    if mode == "set_cc_team":
        Designation.objects.all().update(created_by="cc team")
    if mode == "update_dob_in_profile":
        update_dob_in_profile()
    if mode == "update_total_time":
        update_total_time()
    if mode == "custom_cr_sh":
        custom_cr_sh()
    if mode == "test_db":
        profiles = Profile.objects.all()[0]
        print(profiles)
    if mode == "create_in_qms":
        profile = Profile.objects.filter(emp_id=10326).last()
        if check_server_status(settings.QMS_URL):
            update_profile_in_qms(profile.emp_id)
    if mode == "create_all_profiles_in_qms":
        update_qms_profiles()
    if mode == "decrypt_payslip":
        decrypt_payslip_encrypted("aaaaaa.xlsx")

    if mode == "migrate_payslip":
        migrate_payslip()

    if mode == "set_ams_training":
        date1 = datetime.today().replace(year=2023, month=11, day=1).date()
        attendances = Attendance.objects.filter(date__gte=date1, is_training=True, status="Present")
        attendances.update(status="Training")
    if mode == "wk_sch":
        Attendance.objects.filter(status="Working").update(status="Scheduled")

    if mode == "correct_lwd":
        correct_lwd()

    if mode == "update_applied_leaves":
        update_applied_leaves()

    if mode == "import_vp":
        import_variable_pay("variable_pay.xlsx")
    if mode == "import_lbh":
        import_leave_balance_history('leave_balance_history.xlsx')
    if mode == "update_lbh_date":
        update_leave_balance_history('LeaveBalanceHistory-2023-12-07.xlsx')

    if mode == "split_address":
        split_address()

    if mode == "manual_add_lb":
        # TODO Update the date in LeaveBalanceHistory table from auto_add_now
        date_now = tz.localize(datetime.now().replace(year=2023, month=11, day=1))
        update_leave_balance = False
        attendance_view.compute_leave_balance(non_attrition_profiles(),
                                              date_now=date_now,
                                              update_leave_balance=update_leave_balance)
        attendance_view.compute_leave_balance(get_attrition_profiles_from_1st_last_month(), date_now=date_now,
                                              update_leave_balance=update_leave_balance)

    if mode == "add_lb":
        # for all the months
        attendance_view.compute_leave_balance(non_attrition_profiles())
        attendance_view.compute_leave_balance(get_attrition_profiles_from_1st_last_month())

    if mode == "import_pay_slip":
        import_pay_slip('Payroll-2024-01-10.xlsx')

    # if mode == "check_lb_and_applied_leaves":
    #     # For new year changes
    #     logger.info("updating year lb")
    #     check_and_deduct_lb_and_applied_leaves()
    #     add_new_year_sl(non_attrition_profiles())

    if mode == "lb_for_new_year":
        logger.info("running lb_for_new_year")
        # attendance_view.compute_leave_balance(Profile.objects.filter(emp_id=9194))
        # # attendance_view.compute_leave_balance(get_attrition_profiles_from_1st_last_month())
        # check_and_deduct_lb_and_applied_leaves()
        # add_new_year_sl(Profile.objects.filter(emp_id=9194))

        attendance_view.compute_leave_balance(non_attrition_profiles())
        attendance_view.compute_leave_balance(get_attrition_profiles_from_1st_last_month())
        check_and_deduct_lb_and_applied_leaves()
        add_new_year_sl(non_attrition_profiles())

    if mode == "corrections":
        Application.objects.filter(branch="HBR LAYOUT").update(branch="Hbr Layout")

    if mode == "import_sa":
        import_supplier_administration()

    if mode == "import_po":
        import_po()

    if mode == "update_application_status":
        candidates = Candidate.objects.filter(current_application__isnull=False)
        for candidate in candidates:
            candidate.all_applications.update(status=False)
            candidate.current_application.status = True
            candidate.current_application.save()
    if mode == "update_attrition_profiles":
        profiles = Profile.objects.filter(status="Attrition", onboard__isnull=False)
        for profile in profiles:
            candidate = profile.onboard.candidate
            candidate.all_applications.update(status=False)
            candidate.status = "Attrition"
            candidate.is_employee = False
            candidate.save()

    if mode == "import_name":
        data = read_file_format_output("Profile-2024-02-12.xlsx")
        count = 0
        emp_ids = data["emp_id"]
        for emp_id in emp_ids:
            profile = Profile.objects.filter(emp_id=emp_id, is_superuser=False).last()
            if not profile:
                count += 1
                continue
            profile.first_name = data["first_name"][count]
            profile.middle_name = data["middle_name"][count]
            profile.last_name = data["last_name"][count]
            profile.full_name = data["full_name"][count]
            profile.save()
            count += 1

    if mode == "create_partb_partc_parameters":
        param_list = []
        for i in range(1, 23):
            param_list.append(Parameter(name="b" + str(i), type="B"))
        for i in range(1, 6):
            param_list.append(Parameter(name="c" + str(i), type="C"))
        Parameter.objects.bulk_create(param_list)

    if mode == "create_category_group_mapping":
        hrms_groups = HrmsPermissionGroup.objects.filter(~Q(name="TL"))
        tl_am_mgr_mgmt = HrmsPermissionGroup.objects.filter(name="TL").last()
        for hgroup in hrms_groups:
            for perm in tl_am_mgr_mgmt.permissions.all():
                if perm in hgroup.permissions.all():
                    hgroup.permissions.remove(perm)
            hgroup.category_group.add(hgroup.category)

    if mode == "calc_score":
        score_rating = [(10, 2), (15, 3), (5, 4), (5, 5), (15, 5)]
        score_rating.append(((100 - sum([float(i) for i, j in score_rating])), 3))
        total_score = 0
        for i, j in score_rating:
            total_score += round((float(i) * j) / 100, 2)
        parta = [3, 4, 4, 4, 5, 5, 3, 4, 5, 4, 5, 5, 3, 4, 5, 4, 5, 5, 3, 4, 5, 4]
        partc = [3, 5, 3, 2, 1]
        print(round((total_score * 70 + round((sum(parta) / 22), 2) * 20 + round((sum(partc) / 5), 2) * 10) / 100, 2))

    if mode == "reject_unattended_leaves":
        today = datetime.today().date()
        prev_2nd_month_last_day = (today.replace(day=1) - timedelta(days=1)).replace(day=1) - timedelta(days=1)
        first_day_of_2nd_month = prev_2nd_month_last_day.replace(day=1)
        leave_list = []
        print("came")
        if today.month == 2:
            print("no action taken for the last year december")
            return
        leaves = Leave.objects.filter(status="Pending", start_date__gte=first_day_of_2nd_month,
                                      start_date__lte=prev_2nd_month_last_day)
        for leave in leaves:
            leave_balance = LeaveBalance.objects.get(profile=leave.profile)
            print("came here")
            return_back_the_leaves_to_employee(leave, leave_balance, "Leave Rejected")
            leave.status = "Rejected"
            if leave.manager_comments is None:
                leave.manager_comments = "Leave rejected by system, RM2 no action taken"
            elif leave.tl_comments is None:
                leave.tl_comments = "Leave rejected by system, RM1 no action taken"
            leave.updated_at = tz.localize(datetime.now())
            leave_list.append(leave)
        if len(leave_list) > 0:
            print("came here 2")
            Leave.objects.bulk_update(leave_list, ["status", "manager_comments", "tl_comments", "updated_at"])
    if mode == "update_vpay_status":
        vpays = VariablePay.objects.all().update(status="Approved")

    if mode == "get_non_onboard":
        profiles = Profile.objects.filter(status="Active", onboard__isnull=True)
        start_no = 2000
        columns = ["emp_id", "aadhaar_no", "mobile", "email", "total_fixed_gross", "pay_grade", "opting_for_pf",
                   "pf_emp", "pf_emr", "basic", "hra", "oa", "ta", "tel_allowance", "medic_a",
                   "health_insurance_deduction", "prof_tax", "net_salary"]
        emp_data = []
        today = datetime.now().date()
        prev_month = today.replace(day=1) - timedelta(days=1)
        prev_to_prev_month = prev_month.replace(day=1) - timedelta(days=1)
        for profile in profiles:
            mobile = "{:<010}".format(start_no)
            aadhaar = "{:<012}".format(start_no)
            email = "".join(generate_unique_letters()) + "@2test2.com"
            payroll = Payroll.objects.filter(profile=profile, month_number=prev_month.month,
                                             year=prev_month.year).first()
            if payroll is None:
                payroll = Payroll.objects.filter(profile=profile, month_number=prev_to_prev_month.month,
                                                 year=prev_to_prev_month.year).first()
                if payroll is None:
                    print(f"profile {profile}")
                    continue
            print(f"{mobile},{aadhaar},{email}")
            opting_for_pf = 1 if payroll.pf_emp > 0 else 0
            emp_data.append(
                [profile.emp_id, aadhaar, mobile, email, payroll.total_fixed_gross, payroll.pay_grade, opting_for_pf,
                 payroll.pf_emp, payroll.pf_emr, payroll.basic, payroll.hra, payroll.oa, payroll.ta,
                 payroll.tel_allowance, payroll.medical_a, payroll.health_insurance_deduction, payroll.pt,
                 payroll.net_salary])
            start_no += 1

        with open("pending_prof_onboard.csv", 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(columns)
            for data in emp_data:
                writer.writerow(data)
    if mode == "compare_payroll":
        # columns, _ = get_column_names(Payroll)
        sandwich_count = models.IntegerField(default=0)
        sandwich_dates = models.CharField(max_length=255, blank=True, null=True)

        columns = ["id", "emp_id", "year", "month", "month_number", "payment_mode", "pay_check", "new_qa", "emp_name",
                   "doj", "lwd", "status", "gender", "process", "designation", "department", "bank_name",
                   "name_as_per_bank_acc", "account_number", "ifsc_code", "pay_grade", "da", "new_da", "uan", "pf_no",
                   "esic_no", "present", "wo", "el", "sl", "cl", "ml", "co", "trng", "hd", "nh", "absent", "att", "na",
                   "paid_days", "calendar_days", "total_days", "lop_days", "fixed_gross", "basic", "ta", "hra", "oa",
                   "tel_allowance", "medical_allowance", "total_fixed_gross", "b_plus_d", "el_accrual",
                   "new_basic_salary", "new_oa", "new_ta", "new_hra", "new_medical_a", "pf_gross",
                   "sandwich_count", "sandwich_dates", "statutory_bonus",
                   "variable_pay", "process_incentive", "client_incentive", "bb_rent", "overtime",
                   "night_shift_allowance", "referral", "arrears", "attendance_bonus", "total_gross_salary", "pf_emp",
                   "pf_emr", "esi", "pt", "tds", "health_insurance_deduction", "sodexo_deduction", "cab_deduction",
                   "other_deduction", "notice_period_recovery", "total_ded", "net_salary", "created_at", "updated_at",
                   "profile_id", "branch", "welfare_fund"]
        # columns = ["emp_id", "year", "month", "total_gross_salary", "variable_pay", "net_salary", "actual_net_salary"]
        pay_rolls = Payroll.objects.filter(year=2024, month_number=4)
        pay_rolls2 = Payroll.objects.filter(year=2024, month_number=5)
        pay_data = []
        for pay in pay_rolls:
            salary = Salary.objects.filter(profile=pay.profile).first()
            if salary is None:
                continue
            actual_net_salary = get_decrypted_value(salary.net_salary)
            if abs(int(actual_net_salary) - int(pay.net_salary)) > 1:
                # pay_data.append(
                #     [pay.profile.emp_id, pay.year, pay.month_number, pay.total_gross_salary, pay.variable_pay,
                #      pay.net_salary, actual_net_salary])
                pay2 = pay_rolls2.filter(profile=pay.profile).first()
                # if int(pay2.variable_pay) == int(pay.variable_pay):
                #     continue

                pay_data.append([getattr(pay, col) for col in columns])
                pay_data.append([getattr(pay2, col) for col in columns])

        with open("compare_payroll.csv", 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(columns)
            for data in pay_data:
                writer.writerow(data)

    if mode == "drop_last_month_payroll":
        today = datetime.today().date()
        prev_month = today.replace(day=1) - timedelta(days=1)
        Payroll.objects.filter(year=prev_month.year, month_number=prev_month.month).delete()

        last_month_last_day = today.replace(day=1) - timedelta(days=1)

        attendances = Attendance.objects.filter(date__gte=last_month_last_day.replace(day=1),
                                                date__lte=last_month_last_day, is_sandwich=True)
        attendances.update(status="Week Off", is_sandwich=False)

    if mode == "update_ctc":
        objs = Onboard.objects.all()
        for ob in objs:
            ob.ctc = ob.total_fixed_gross * 12
            ob.salary = ob.appointment_letter.salary
            ob.save()

    if mode == "separate_salary":
        objs = Onboard.objects.all()
        cr_salary_list = []
        from onboarding.serializer import UpdateSalarySerializer
        for obj in objs:
            salary = Salary.objects.filter(candidate=obj.candidate).first()
            if salary is None:
                salary_data = update_request(obj.__dict__, candidate=obj.candidate.id,
                                             updated_by=obj.onboarded_by.id if obj.onboarded_by else None,
                                             last_update_reason="Onboard")
                serializer = UpdateSalarySerializer(data=salary_data)
                if not serializer.is_valid():
                    return_error_response(serializer, raise_exception=True)
                    break
                salary = serializer.create(serializer.validated_data)
                obj.appointment_letter.salary = salary
                obj.salary = salary
                obj.appointment_letter.save()
                obj.save()

    if mode == "view_all_get_apis":
        from django.urls import URLResolver, URLPattern
        import importlib
        from rest_framework.views import APIView

        def get_get_urlpatterns(urlpatterns, parent_pattern=None, parent_module=None):
            get_urlpatterns = []
            for pattern in urlpatterns:
                if isinstance(pattern, URLPattern):
                    # Check if the pattern allows GET method
                    # Check if the view associated with the URL pattern allows GET method
                    if hasattr(pattern.callback, 'view_class') and issubclass(pattern.callback.view_class, APIView):
                        # Get the HTTP methods allowed for the view
                        allowed_methods = pattern.callback.view_class().allowed_methods
                        # Check if GET method is allowed
                        if 'get' in map(str.lower, allowed_methods):
                            get_urlpatterns.append(
                                (pattern.pattern, parent_pattern.pattern.regex.pattern[1::] if parent_pattern else ""))
                elif isinstance(pattern, URLResolver):
                    # Recursively call the function to handle include() patterns
                    get_urlpatterns.extend(
                        get_get_urlpatterns(pattern.url_patterns, pattern, parent_module=parent_module))
                elif hasattr(pattern, 'urlconf_name'):
                    # If the pattern is an included URLconf, get the module name
                    urlconf_module = importlib.import_module(pattern.urlconf_name)
                    get_urlpatterns.extend(
                        get_get_urlpatterns(urlconf_module.urlpatterns, parent_pattern, pattern.urlconf_name))

            return get_urlpatterns

        # Get all urlpatterns from the project's main urls.py
        root_urlconf = importlib.import_module(settings.ROOT_URLCONF)
        get_urlpatterns = get_get_urlpatterns(root_urlconf.urlpatterns)

        # Print view name, endpoint, and module name for each GET URL pattern
        for view_name, endpoint in get_urlpatterns:
            print(f"View Name: {view_name}, Endpoint: {endpoint}")

    if mode == "update_test_passwords":
        if settings.is_development:
            user_credentials = [{"username": i, "password": f"Ecpl@{i}"} for i in settings.TEST_ACCOUNTS]

            # Change passwords for each user
            for user in user_credentials:
                change_password(user["username"], user["password"])

    if mode == "link_application_candidate":
        candidates = Candidate.objects.filter(current_application__isnull=False)
        for candidate in candidates:
            candidate.all_applications.all().update(candidate=candidate)

    if mode == "add_answers_score":
        add_score()

    if mode == "cron_auto_approve_attendance":
        today = datetime.today().date()
        if today.day != 1:
            logger.info("today is not first")
            return
        first = today.replace(day=1)
        last_month_last_day = today.replace(day=1) - timedelta(days=1)
        if first.strftime("%A") == "Saturday":
            days = [last_month_last_day]
        elif first.strftime("%A") == "Sunday":
            days = [last_month_last_day, last_month_last_day - timedelta(days=1)]
        else:
            logger.info("first it not saturday or sunday")
            return

        attendances = Attendance.objects.filter(date__in=days, status="Scheduled", total_time__isnull=False)
        if len(attendances) == 0:
            logger.info("there are no attendances to approve")
        for attendance in attendances:
            check_if_day_is_endofmonth_for_auto_approval_custom_seed_one(attendance)
    if mode=="create_or_update_profile":
        emp_ids=Profile.objects.filter(is_superuser=False).values_list("emp_id",flat=True)
        for emp_id in emp_ids:
            update_profile_in_qms3(emp_id)
    if mode=="update_profile_dependencies":
        logger.info(f"-----start Profile dependency update or create----")
        create_update_dependencies()
    if mode=="ijp_selection_conversion":
        total_ijps=IJP.objects.all()
        for ijp in total_ijps:
            ijp.selection_criteria=json.dumps(ijp.selection_criteria.split(","))
            ijp.selection_process=json.dumps(ijp.selection_process.split(","))
            ijp.save()






def generate_unique_letters():
    # Generate a list of lowercase letters
    letters = list(string.ascii_lowercase)
    # Shuffle the letters
    random.shuffle(letters)
    # Take the first 20 unique letters
    unique_letters = letters[:20]
    return unique_letters
